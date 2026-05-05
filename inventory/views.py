from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import StockOrder, StockOrderItem
# Save Stock Order via AJAX
@csrf_exempt
@login_required
def save_stock_order(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode())
        manufacturer = data.get('manufacturer')
        order_date = data.get('order_date')
        items = data.get('items', [])
        order = StockOrder.objects.create(
            manufacturer=manufacturer,
            order_date=order_date,
            created_by=request.user if request.user.is_authenticated else None
        )
        for item in items:
            StockOrderItem.objects.create(
                order=order,
                kulfi_name=item.get('name'),
                lot=item.get('lot', 0),
                quantity=item.get('qty', 0)
            )
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)
from django.views.decorators.http import require_POST
from django.core.mail import send_mail
import json
# ==================== AUTHENTICATION ====================

# Forgot Password Email Handler
@require_POST
def send_forgot_password_email(request):
    try:
        data = json.loads(request.body.decode()) if request.body else {}
    except Exception:
        data = {}
    user_email = 'jsdansari@gmail.com'
    user_ip = request.META.get('REMOTE_ADDR', 'Unknown IP')
    ua = request.META.get('HTTP_USER_AGENT', 'Unknown')
    send_mail(
        subject='Indian Kulfi Inventory: Forgot Password Request',
        message=f"A user has requested a password reset from the login page.\n\nIP: {user_ip}\nUser-Agent: {ua}",
        from_email=None,
        recipient_list=[user_email],
        fail_silently=False,
    )
    return JsonResponse({'message': 'Password reset request sent. The admin will contact you.'})
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count, Q, F, DecimalField, OuterRef, Subquery, Case, When, IntegerField
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from datetime import datetime, timedelta, date
from decimal import Decimal
from urllib.parse import urlencode
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse
from django.views.decorators.http import require_POST
from functools import wraps
import csv
import io
import os
import re
from collections import OrderedDict, defaultdict

from .models import Product, Inventory, Sales, SalesStockTaken, OperationsExpense, OperationsIncome, DailySalesReport, WeeklyReport, ProfitReport, SalesCountDraft, ExpenseDetailOption
from .forms import ProductForm, SalesForm, DateRangeForm, OperationsExpenseForm, OperationsIncomeForm, UserManagementForm


# Fixed Indian Kulfi costs used in Quick Inventory Entry when manufacturer is Indian Kulfi.
IK_QUICK_ENTRY_COST_BY_NAME = {
    'malai': Decimal('24.17'),
    'kesar badam': Decimal('24.17'),
    'kesar pista': Decimal('24.17'),
    'pista badam': Decimal('26.67'),
    'chocolate': Decimal('26.67'),
    'strawberry': Decimal('24.17'),
    'mango malai': Decimal('24.17'),
    'dry fruit': Decimal('26.67'),
    'butterscotch': Decimal('26.67'),
    'rose': Decimal('26.67'),
    'blackcurrent': Decimal('26.67'),
    'blackcurrant': Decimal('26.67'),
    'caramel coffee': Decimal('26.67'),
    'coconut': Decimal('24.17'),
    'elaichi': Decimal('24.17'),
    'litchi': Decimal('24.17'),
    'kesar kajoor': Decimal('24.17'),
    'guava': Decimal('26.67'),
    'paan': Decimal('26.67'),
}

# Fixed Kulfi Corner costs used in Quick Inventory Entry when manufacturer is Kulfi Corner.
KC_QUICK_ENTRY_COST_BY_NAME = {
    'malai': Decimal('28.00'),
    'kesar badam': Decimal('28.00'),
    'kesar pista': Decimal('28.00'),
    'pista badam': Decimal('28.00'),
    'chocolate': Decimal('28.00'),
    'strawberry': Decimal('26.00'),
    'mango malai': Decimal('26.00'),
    'dry fruit': Decimal('28.00'),
    'butterscotch': Decimal('28.00'),
    'rose': Decimal('28.00'),
    'blackcurrent': Decimal('28.00'),
    'blackcurrant': Decimal('28.00'),
    'caramel coffee': Decimal('28.00'),
    'coconut': Decimal('26.00'),
    'elaichi': Decimal('26.00'),
    'litchi': Decimal('26.00'),
    'kesar kajoor': Decimal('26.00'),
    'guava': Decimal('28.00'),
    'paan': Decimal('28.00'),
}

# Cost price → manufacturer lookup (covers both IK and KC known price points).
COST_PRICE_MANUFACTURER_MAP = {
    Decimal('24.17'): 'Indian Kulfi',
    Decimal('26.67'): 'Indian Kulfi',
    Decimal('26.00'): 'Kulfi Corner',
    Decimal('28.00'): 'Kulfi Corner',
}


def _identify_manufacturer_from_cost(cost_price):
    """Return manufacturer name derived from cost price, defaulting to 'Kulfi Corner'."""
    return COST_PRICE_MANUFACTURER_MAP.get(Decimal(str(cost_price)), 'Kulfi Corner')


def _extract_manufacturer_from_notes(notes):
    """Return the manufacturer name stamped in an inventory movement's notes, or None."""
    if not notes:
        return None
    for segment in notes.split('|'):
        segment = segment.strip()
        if segment.startswith('Manufacturer:'):
            value = segment[len('Manufacturer:'):].strip()
            return value if value else None
    return None


def _get_report_logo_path():
    """Return the first available local logo path suitable for ReportLab PDFs."""
    candidates = [
        os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.png'),
        os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpg'),
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return None


def admin_only_view(view_func):
    """Restrict view access to staff users only."""
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_staff:
            messages.error(request, 'You can access only the Sales module with this account.')
            return redirect('quick_sales_entry')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# ==================== AUTHENTICATION ====================

def login_view(request):
    """User login view"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            if user.is_staff:
                return redirect('dashboard')
            return redirect('quick_sales_entry')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'inventory/login.html')

@login_required
def logout_view(request):
    """User logout view"""
    logout(request)
    return redirect('login')

# ==================== DASHBOARD ====================

@login_required
def dashboard(request):
    """Main dashboard showing today's sales, revenue, and low stock alerts"""
    today = timezone.now().date()
    
    # Today's sales
    today_sales = Sales.objects.filter(sale_date=today)
    total_today_sales = today_sales.count()
    total_today_revenue = today_sales.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']
    
    # Gross profit from sales only.
    total_today_profit = sum(sale.get_profit() for sale in today_sales)
    total_today_operation_cost = OperationsExpense.objects.filter(operation_date=today).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    total_today_net_profit = total_today_profit - total_today_operation_cost
    
    # Total stock across active products only
    total_stock = Product.objects.filter(is_active=True).aggregate(
        total=Coalesce(
            Sum(
                Case(
                    When(current_stock__gt=0, then='current_stock'),
                    default=0,
                    output_field=IntegerField(),
                )
            ),
            0,
            output_field=DecimalField(),
        )
    )['total']

    # Low stock alerts
    low_stock_products = Product.objects.filter(
        current_stock__lte=F('reorder_level'),
        is_active=True
    )
    
    # Weekly sales trend (last 7 days)
    last_7_days_sales = []
    for i in range(6, -1, -1):
        date_temp = today - timedelta(days=i)
        sales_count = Sales.objects.filter(sale_date=date_temp).count()
        revenue = Sales.objects.filter(sale_date=date_temp).aggregate(
            total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
        )['total']
        last_7_days_sales.append({
            'date': date_temp.strftime('%m/%d'),
            'sales': sales_count,
            'revenue': float(revenue)
        })
    
    # Top products (by sales count)
    top_products = Product.objects.annotate(
        sale_count=Count('sales')
    ).order_by('-sale_count')[:5]
    
    context = {
        'total_stock': total_stock,
        'total_today_sales': total_today_sales,
        'total_today_revenue': total_today_revenue,
        'total_today_profit': total_today_profit,
        'total_today_operation_cost': total_today_operation_cost,
        'total_today_net_profit': total_today_net_profit,
        'low_stock_count': low_stock_products.count(),
        'low_stock_products': low_stock_products[:5],
        'last_7_days_sales': last_7_days_sales,
        'top_products': top_products,
    }
    
    return render(request, 'inventory/dashboard.html', context)

# ==================== INDIAN KULFI PRODUCTS MODULE ====================

@login_required
@permission_required('inventory.add_product', raise_exception=True)
def add_product(request):
    """Add new product to inventory"""
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            messages.success(request, f'Product "{product.name}" added successfully!')
            return redirect('product_list')
    else:
        form = ProductForm()
    
    context = {'form': form, 'title': 'Add New Product'}
    return render(request, 'inventory/add_product.html', context)


@login_required
def product_list(request):
    """List all products with management options"""
    products = Product.objects.filter(is_active=True).order_by('sku')
    
    # Pagination
    paginator = Paginator(products, 25)
    page_number = request.GET.get('page', 1)
    try:
        paginated_products = paginator.page(page_number)
    except PageNotAnInteger:
        paginated_products = paginator.page(1)
    except EmptyPage:
        paginated_products = paginator.page(paginator.num_pages)
    
    context = {
        'products': paginated_products,
        'paginator': paginator,
        'page_number': paginated_products.number,
    }
    return render(request, 'inventory/product_list.html', context)

@login_required
@permission_required('inventory.change_product', raise_exception=True)
def edit_product(request, product_id):
    """Edit existing product"""
    product = get_object_or_404(Product, pk=product_id)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            product = form.save()
            messages.success(request, f'Product "{product.name}" updated successfully!')
            return redirect('inventory_list')
    else:
        form = ProductForm(instance=product)
    
    context = {'form': form, 'product': product, 'title': f'Edit {product.name}'}
    return render(request, 'inventory/add_product.html', context)

@login_required
@permission_required('inventory.delete_product', raise_exception=True)
def delete_product(request, product_id):
    """Soft delete a product by marking it inactive"""
    product = get_object_or_404(Product, pk=product_id)

    if request.method == 'POST':
        product.is_active = False
        product.save()
        messages.success(request, f'Product "{product.name}" deleted successfully.')
        return redirect('inventory_list')

    context = {'product': product}
    return render(request, 'inventory/confirm_delete_product.html', context)


@login_required
@permission_required('inventory.delete_product', raise_exception=True)
def trash_list(request):
    """List soft-deleted products (trash)"""
    trashed_products = Product.objects.filter(is_active=False).order_by('name')
    context = {'products': trashed_products}
    return render(request, 'inventory/trash_list.html', context)


@login_required
@permission_required('inventory.change_product', raise_exception=True)
def restore_product(request, product_id):
    """Restore soft-deleted product"""
    product = get_object_or_404(Product, pk=product_id, is_active=False)
    product.is_active = True
    product.save()
    messages.success(request, f'Product "{product.name}" restored successfully.')
    return redirect('inventory_trash')


@login_required
@permission_required('inventory.delete_product', raise_exception=True)
def hard_delete_product(request, product_id):
    """Permanently hard delete product"""
    product = get_object_or_404(Product, pk=product_id, is_active=False)
    product.delete()
    messages.success(request, f'Product "{product.name}" permanently deleted.')
    return redirect('inventory_trash')


# ==================== INVENTORY MODULE ====================

def calculate_stock_as_of_date(product, selected_date):
    """Calculate product stock as of a specific date using inventory movements."""
    last_adjustment = Inventory.objects.filter(
        product=product,
        movement_type='ADJUSTMENT',
        movement_date__lte=selected_date
    ).order_by('-movement_date', '-created_at').first()

    if last_adjustment:
        base_date = last_adjustment.movement_date
        base_qty = last_adjustment.quantity
        base_created_at = last_adjustment.created_at
        movement_sums = Inventory.objects.filter(
            product=product,
            movement_date__lte=selected_date
        ).filter(
            Q(movement_date__gt=base_date) |
            Q(movement_date=base_date, created_at__gt=base_created_at)
        ).aggregate(
            in_total=Coalesce(Sum('quantity', filter=Q(movement_type='IN')), 0),
            out_total=Coalesce(Sum('quantity', filter=Q(movement_type='OUT')), 0)
        )

        # Sales reduce stock and must be included in historical stock reconstruction.
        sales_total = Sales.objects.filter(
            product=product,
            sale_date__lte=selected_date
        ).filter(
            Q(sale_date__gt=base_date) |
            Q(sale_date=base_date, created_at__gt=base_created_at)
        ).aggregate(
            total=Coalesce(Sum('quantity'), 0)
        )['total']

        calculated = base_qty + movement_sums['in_total'] - movement_sums['out_total'] - sales_total
        return max(0, calculated)

    movement_sums = Inventory.objects.filter(
        product=product,
        movement_date__lte=selected_date
    ).aggregate(
        in_total=Coalesce(Sum('quantity', filter=Q(movement_type='IN')), 0),
        out_total=Coalesce(Sum('quantity', filter=Q(movement_type='OUT')), 0)
    )

    sales_total = Sales.objects.filter(
        product=product,
        sale_date__lte=selected_date
    ).aggregate(
        total=Coalesce(Sum('quantity'), 0)
    )['total']

    calculated = movement_sums['in_total'] - movement_sums['out_total'] - sales_total
    return max(0, calculated)


def get_stock_as_of_date_map(products, selected_date):
    """Return a {product_id: stock_as_of_date} map for the provided products."""
    product_ids = [product.id for product in products]
    if not product_ids:
        return {}

    movement_totals_map = {
        row['product_id']: row
        for row in Inventory.objects.filter(
            product_id__in=product_ids,
            movement_date__lte=selected_date
        ).values('product_id').annotate(
            in_total=Coalesce(Sum('quantity', filter=Q(movement_type='IN')), 0),
            out_total=Coalesce(Sum('quantity', filter=Q(movement_type='OUT')), 0)
        )
    }

    sales_totals_map = {
        row['product_id']: row['total']
        for row in Sales.objects.filter(
            product_id__in=product_ids,
            sale_date__lte=selected_date
        ).values('product_id').annotate(
            total=Coalesce(Sum('quantity'), 0)
        )
    }

    last_adjustment_query = Inventory.objects.filter(
        product=OuterRef('pk'),
        movement_type='ADJUSTMENT',
        movement_date__lte=selected_date
    ).order_by('-movement_date', '-created_at')

    last_adjustments_map = {}
    for row in Product.objects.filter(id__in=product_ids).annotate(
        last_adjustment_date=Subquery(last_adjustment_query.values('movement_date')[:1]),
        last_adjustment_qty=Subquery(last_adjustment_query.values('quantity')[:1]),
        last_adjustment_created_at=Subquery(last_adjustment_query.values('created_at')[:1]),
    ).values('id', 'last_adjustment_date', 'last_adjustment_qty', 'last_adjustment_created_at'):
        if row['last_adjustment_date'] is not None:
            last_adjustments_map[row['id']] = (
                row['last_adjustment_date'],
                row['last_adjustment_qty'],
                row['last_adjustment_created_at'],
            )

    stock_map = {}
    for product_id in product_ids:
        if product_id in last_adjustments_map:
            base_date, base_qty, base_created_at = last_adjustments_map[product_id]
            movement_after = Inventory.objects.filter(
                product_id=product_id,
                movement_date__lte=selected_date
            ).filter(
                Q(movement_date__gt=base_date) |
                Q(movement_date=base_date, created_at__gt=base_created_at)
            ).aggregate(
                in_total=Coalesce(Sum('quantity', filter=Q(movement_type='IN')), 0),
                out_total=Coalesce(Sum('quantity', filter=Q(movement_type='OUT')), 0)
            )
            sales_after = Sales.objects.filter(
                product_id=product_id,
                sale_date__lte=selected_date
            ).filter(
                Q(sale_date__gt=base_date) |
                Q(sale_date=base_date, created_at__gt=base_created_at)
            ).aggregate(
                total=Coalesce(Sum('quantity'), 0)
            )['total']
            calculated = base_qty + movement_after['in_total'] - movement_after['out_total'] - sales_after
            stock_map[product_id] = max(0, calculated)
            continue

        movement_totals = movement_totals_map.get(product_id, {'in_total': 0, 'out_total': 0})
        sales_total = sales_totals_map.get(product_id, 0)
        calculated = movement_totals['in_total'] - movement_totals['out_total'] - sales_total
        stock_map[product_id] = max(0, calculated)

    return stock_map


@login_required
def inventory_list(request):
    """View inventory with current or date-based stock levels."""
    clear_filters = request.GET.get('clear') == '1'
    reset_filters = request.GET.get('reset') == '1'
    show_results = not clear_filters

    if clear_filters or reset_filters:
        search_query = ''
        category_filter = []
        status_filter = ''
        movement_filter = ''
        as_of_date = ''
        sort_by = 'sku'
        per_page = 50

        request.session.pop('inventory_search_query', None)
        request.session.pop('inventory_category_filter', None)
        request.session.pop('inventory_status_filter', None)
        request.session.pop('inventory_movement_filter', None)
        request.session.pop('inventory_as_of_date', None)
        request.session.pop('inventory_sort_by', None)
        request.session.pop('inventory_per_page', None)
    else:
        search_query = request.GET.get('search', request.session.get('inventory_search_query', ''))
        category_filter = request.GET.getlist('category') or request.session.get('inventory_category_filter', [])
        status_filter = request.GET.get('status', request.session.get('inventory_status_filter', ''))
        movement_filter = request.GET.get('movement_type', request.session.get('inventory_movement_filter', ''))
        as_of_date = request.GET.get('as_of_date', request.session.get('inventory_as_of_date', ''))
        sort_by = request.GET.get('sort', request.session.get('inventory_sort_by', 'sku'))
        per_page = int(request.GET.get('per_page', request.session.get('inventory_per_page', 50)))

    request.session['inventory_search_query'] = search_query
    request.session['inventory_category_filter'] = category_filter
    request.session['inventory_status_filter'] = status_filter
    request.session['inventory_movement_filter'] = movement_filter
    request.session['inventory_as_of_date'] = as_of_date
    request.session['inventory_sort_by'] = sort_by
    request.session['inventory_per_page'] = per_page

    selected_date = None
    if as_of_date:
        try:
            selected_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
            as_of_date = selected_date.isoformat()

    generated_at = timezone.localtime()
    if selected_date:
        stock_as_of_label = f"{selected_date.strftime('%Y-%m-%d')} 23:59:59"
    else:
        stock_as_of_label = generated_at.strftime('%Y-%m-%d %H:%M:%S')

    products = Product.objects.filter(is_active=True)

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(category__icontains=search_query)
        )

    if category_filter:
        products = products.filter(category__in=category_filter)

    categories = ['Indian Kulfi']

    sort_options = {
        'name': 'name',
        '-name': '-name',
        'sku': 'sku',
        '-sku': '-sku',
        'current_stock': 'display_stock',
        '-current_stock': '-display_stock',
        'reorder_level': 'reorder_level',
        '-reorder_level': '-reorder_level',
        'cost_price': 'cost_price',
        '-cost_price': '-cost_price',
        'selling_price': 'selling_price',
        '-selling_price': '-selling_price',
    }

    products_ordered = list(products.order_by('sku'))
    stock_map = get_stock_as_of_date_map(products_ordered, selected_date) if selected_date else {}

    # Ensure all products have display_stock set
    for product in products_ordered:
        if selected_date:
            product.display_stock = max(0, stock_map.get(product.id, 0))
        else:
            product.display_stock = max(0, product.current_stock)
        # Remove 'KC' from product name for display
        if 'KC' in product.name:
            product.name = product.name.replace('KC', '').strip()

    movement_product_ids = set()
    if movement_filter:
        movement_query = Inventory.objects.filter(
            product_id__in=[product.id for product in products_ordered],
            movement_type=movement_filter
        )
        if selected_date:
            movement_query = movement_query.filter(movement_date=selected_date)
        movement_product_ids = set(
            movement_query.values_list('product_id', flat=True).distinct()
        )

    product_list = []
    for product in products_ordered:
        display_stock = product.display_stock

        movement_exists = True
        if movement_filter:
            movement_exists = product.id in movement_product_ids

        if not movement_exists:
            continue

        if status_filter == 'low_stock' and display_stock > product.reorder_level:
            continue
        if status_filter == 'in_stock' and display_stock <= product.reorder_level:
            continue

        product_list.append(product)

    sort_key = sort_options.get(sort_by, 'sku')
    reverse_sort = sort_key.startswith('-')
    sort_attr = sort_key[1:] if reverse_sort else sort_key

    # Custom display order when sorting by SKU (default view)
    _KULFI_SKU_ORDER = [
        'IK0001', 'IK0004', 'IK0005', 'IK0002', 'IK0003', 'IK0006',
        'IK0008', 'IK0011', 'IK0015', 'IK0012', 'IK0010', 'IK0007',
        'IK0009', 'IK0013', 'IK0014', 'IK0017', 'IK0018', 'IK0016',
    ]
    if sort_attr == 'sku' and not reverse_sort:
        _sku_pos = {sku: i for i, sku in enumerate(_KULFI_SKU_ORDER)}
        product_list.sort(key=lambda p: _sku_pos.get(p.sku, len(_KULFI_SKU_ORDER)))
    else:
        product_list.sort(key=lambda product: getattr(product, sort_attr), reverse=reverse_sort)

    # Calculate totals from source data with proper display_stock
    # Recalculate to ensure accuracy with filtered product_list
    total_stock = 0
    total_cost_price = Decimal('0.0')
    total_sales_price = Decimal('0.0')
    
    for product in product_list:
        # Use display_stock which is the calculated stock for the selected date
        qty = product.display_stock if hasattr(product, 'display_stock') else (
            stock_map.get(product.id, 0) if selected_date else product.current_stock
        )
        total_stock += qty
        total_cost_price += Decimal(qty) * product.cost_price
        total_sales_price += Decimal(qty) * product.selling_price

    paginator = Paginator(product_list, per_page)
    page_number = request.GET.get('page', 1)
    try:
        paginated_products = paginator.page(page_number)
    except PageNotAnInteger:
        paginated_products = paginator.page(1)
    except EmptyPage:
        paginated_products = paginator.page(paginator.num_pages)

    context = {
        'products': paginated_products,
        'search_query': search_query,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'movement_filter': movement_filter,
        'as_of_date': as_of_date,
        'sort_by': sort_by,
        'per_page': per_page,
        'sort_options': sort_options,
        'categories': categories,
        'paginator': paginator,
        'page_number': paginated_products.number,
        'total_stock': total_stock,
        'total_cost_price': total_cost_price,
        'total_sales_price': total_sales_price,
        'selected_date': selected_date,
        'stock_as_of_label': stock_as_of_label,
        'generated_at': generated_at,
        'show_results': show_results,
    }

    return render(request, 'inventory/inventory_list.html', context)


def _build_inventory_export_context(request):
    """Build filtered inventory data for View Inventory exports."""
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.getlist('category')
    status_filter = request.GET.get('status', '')
    movement_filter = request.GET.get('movement_type', '')
    as_of_date = request.GET.get('as_of_date', '')
    sort_by = request.GET.get('sort', 'sku')

    selected_date = None
    if as_of_date:
        try:
            selected_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
            as_of_date = selected_date.isoformat()

    generated_at = timezone.localtime()
    if selected_date:
        stock_as_of_label = f"{selected_date.strftime('%Y-%m-%d')} 23:59:59"
    else:
        stock_as_of_label = generated_at.strftime('%Y-%m-%d %H:%M:%S')

    products = Product.objects.filter(is_active=True)

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(category__icontains=search_query)
        )

    if category_filter:
        products = products.filter(category__in=category_filter)

    sort_options = {
        'name': 'name',
        '-name': '-name',
        'sku': 'sku',
        '-sku': '-sku',
        'current_stock': 'display_stock',
        '-current_stock': '-display_stock',
        'reorder_level': 'reorder_level',
        '-reorder_level': '-reorder_level',
        'cost_price': 'cost_price',
        '-cost_price': '-cost_price',
        'selling_price': 'selling_price',
        '-selling_price': '-selling_price',
    }

    products_ordered = list(products.order_by('sku'))
    stock_map = get_stock_as_of_date_map(products_ordered, selected_date) if selected_date else {}

    for product in products_ordered:
        if selected_date:
            product.display_stock = max(0, stock_map.get(product.id, 0))
        else:
            product.display_stock = max(0, product.current_stock)

    movement_product_ids = set()
    if movement_filter:
        movement_query = Inventory.objects.filter(
            product_id__in=[product.id for product in products_ordered],
            movement_type=movement_filter
        )
        if selected_date:
            movement_query = movement_query.filter(movement_date=selected_date)
        movement_product_ids = set(
            movement_query.values_list('product_id', flat=True).distinct()
        )

    product_list = []
    for product in products_ordered:
        display_stock = product.display_stock

        movement_exists = True
        if movement_filter:
            movement_exists = product.id in movement_product_ids

        if not movement_exists:
            continue

        if status_filter == 'low_stock' and display_stock > product.reorder_level:
            continue
        if status_filter == 'in_stock' and display_stock <= product.reorder_level:
            continue

        product_list.append(product)

    sort_key = sort_options.get(sort_by, 'sku')
    reverse_sort = sort_key.startswith('-')
    sort_attr = sort_key[1:] if reverse_sort else sort_key

    # Custom display order when sorting by SKU (default view)
    _KULFI_SKU_ORDER = [
        'IK0001', 'IK0004', 'IK0005', 'IK0002', 'IK0003', 'IK0006',
        'IK0008', 'IK0011', 'IK0015', 'IK0012', 'IK0010', 'IK0007',
        'IK0009', 'IK0013', 'IK0014', 'IK0017', 'IK0018', 'IK0016',
    ]
    if sort_attr == 'sku' and not reverse_sort:
        _sku_pos = {sku: i for i, sku in enumerate(_KULFI_SKU_ORDER)}
        product_list.sort(key=lambda p: _sku_pos.get(p.sku, len(_KULFI_SKU_ORDER)))
    else:
        product_list.sort(key=lambda product: getattr(product, sort_attr), reverse=reverse_sort)

    total_stock = 0
    total_cost_price = Decimal('0.0')
    total_sales_price = Decimal('0.0')
    for product in product_list:
        qty = max(0, getattr(product, 'display_stock', product.current_stock))
        total_stock += qty
        total_cost_price += Decimal(qty) * product.cost_price
        total_sales_price += Decimal(qty) * product.selling_price

    return {
        'products': product_list,
        'selected_date': selected_date,
        'as_of_date': as_of_date,
        'stock_as_of_label': stock_as_of_label,
        'generated_at': generated_at,
        'search_query': search_query,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'movement_filter': movement_filter,
        'sort_by': sort_by,
        'total_stock': total_stock,
        'total_cost_price': total_cost_price,
        'total_sales_price': total_sales_price,
    }


@login_required
def print_inventory_html(request):
    """Print-friendly HTML for View Inventory."""
    context = _build_inventory_export_context(request)
    context['now'] = timezone.localtime()
    return render(request, 'inventory/print_inventory.html', context)


@login_required
def print_inventory_pdf(request):
    """Export View Inventory as PDF."""
    context = _build_inventory_export_context(request)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        today_str = timezone.localtime().strftime('%Y-%m-%d')
        response['Content-Disposition'] = f'attachment; filename="inventory_view_{today_str}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.8 * inch, height=0.8 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.1 * inch))

        elements.append(Paragraph('<b>View Inventory Report</b>', styles['Title']))
        elements.append(Spacer(1, 0.15 * inch))
        elements.append(Paragraph(
            (
                f"<b>Stock As Of:</b> {context['stock_as_of_label']} | "
                f"<b>Total Stock:</b> {context['total_stock']} | "
                f"<b>Total Cost Price:</b> Rs.{context['total_cost_price']:.2f} | "
                f"<b>Total Sales Price:</b> Rs.{context['total_sales_price']:.2f}"
            ),
            styles['Normal'],
        ))
        elements.append(Spacer(1, 0.15 * inch))

        data = [['Product', 'SKU', 'Category', 'Stock', 'Reorder', 'Cost Price', 'Selling Price', 'Status']]
        for product in context['products']:
            stock_value = max(0, getattr(product, 'display_stock', product.current_stock))
            status = 'Low Stock' if stock_value <= product.reorder_level else 'In Stock'
            display_name = product.name.replace('KC', '').strip() if 'KC' in product.name else product.name
            data.append([
                display_name,
                product.sku,
                'Indian Kulfi',
                str(stock_value),
                str(product.reorder_level),
                f"Rs.{product.cost_price:.2f}",
                f"Rs.{product.selling_price:.2f}",
                status,
            ])

        table = Table(
            data,
            colWidths=[2.0 * inch, 1.0 * inch, 1.4 * inch, 0.8 * inch, 0.8 * inch, 1.0 * inch, 1.1 * inch, 0.9 * inch],
            repeatRows=1,
        )
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('inventory_list')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('inventory_list')


@login_required
def print_inventory_excel(request):
    """Export View Inventory as Excel."""
    context = _build_inventory_export_context(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory View'

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14

        ws['A1'] = 'VIEW INVENTORY REPORT'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = f"Stock As Of: {context['stock_as_of_label']}"
        ws['A4'] = f"Total Stock: {context['total_stock']}"
        ws['A5'] = f"Total Cost Price: Rs.{context['total_cost_price']:.2f}"
        ws['A6'] = f"Total Sales Price: Rs.{context['total_sales_price']:.2f}"

        headers = ['Product Name', 'SKU', 'Category', 'Stock', 'Reorder Level', 'Cost Price', 'Selling Price', 'Status']
        header_row = 8
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row_number = header_row + 1
        for product in context['products']:
            stock_value = max(0, getattr(product, 'display_stock', product.current_stock))
            status = 'Low Stock' if stock_value <= product.reorder_level else 'In Stock'
            display_name = product.name.replace('KC', '').strip() if 'KC' in product.name else product.name

            ws.cell(row=row_number, column=1).value = display_name
            ws.cell(row=row_number, column=2).value = product.sku
            ws.cell(row=row_number, column=3).value = 'Indian Kulfi'
            ws.cell(row=row_number, column=4).value = int(stock_value)
            ws.cell(row=row_number, column=5).value = int(product.reorder_level)
            ws.cell(row=row_number, column=6).value = float(product.cost_price)
            ws.cell(row=row_number, column=7).value = float(product.selling_price)
            ws.cell(row=row_number, column=8).value = status
            row_number += 1

        for row in range(header_row + 1, row_number):
            ws.cell(row=row, column=6).number_format = '₹#,##0.00'
            ws.cell(row=row, column=7).number_format = '₹#,##0.00'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        today_str = timezone.localtime().strftime('%Y-%m-%d')
        response['Content-Disposition'] = f'attachment; filename="inventory_view_{today_str}.xlsx"'
        wb.save(response)
        return response
    except ImportError as e:
        messages.error(request, f'Excel generation not available. Please install openpyxl: {str(e)}')
        return redirect('inventory_list')
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('inventory_list')


@login_required
def print_inventory_csv(request):
    """Export View Inventory as CSV."""
    context = _build_inventory_export_context(request)

    response = HttpResponse(content_type='text/csv')
    today_str = timezone.localtime().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="inventory_view_{today_str}.csv"'

    writer = csv.writer(response)
    writer.writerow(['View Inventory Report'])
    writer.writerow(['Stock As Of', context['stock_as_of_label']])
    writer.writerow(['Total Stock', context['total_stock']])
    writer.writerow(['Total Cost Price', f"Rs.{context['total_cost_price']:.2f}"])
    writer.writerow(['Total Sales Price', f"Rs.{context['total_sales_price']:.2f}"])
    writer.writerow([])
    writer.writerow(['Product Name', 'SKU', 'Category', 'Stock', 'Reorder Level', 'Cost Price', 'Selling Price', 'Status'])

    for product in context['products']:
        stock_value = max(0, getattr(product, 'display_stock', product.current_stock))
        status = 'Low Stock' if stock_value <= product.reorder_level else 'In Stock'
        display_name = product.name.replace('KC', '').strip() if 'KC' in product.name else product.name
        writer.writerow([
            display_name,
            product.sku,
            'Indian Kulfi',
            stock_value,
            product.reorder_level,
            f"{product.cost_price:.2f}",
            f"{product.selling_price:.2f}",
            status,
        ])

    return response

@login_required
def quick_inventory_entry(request):
    """Quick inventory entry with batch product movements"""
    if request.method == 'POST':
        selected_manufacturer = (request.POST.get('selected_manufacturer') or '').strip()
        products = request.POST.getlist('product[]')
        movement_types = request.POST.getlist('movement_type[]')
        adjustment_modes = request.POST.getlist('adjustment_mode[]')
        quantities = request.POST.getlist('quantity[]')
        quantity_units = request.POST.getlist('quantity_unit[]')
        movement_dates = request.POST.getlist('movement_date[]')

        created_count = 0
        errors = []
        product_count = 0
        total_packs = 0
        total_effective_qty = 0
        overall_total_value = Decimal('0.0')
        recorded_items = []

        for i, product_id in enumerate(products):
            if not product_id:
                continue

            try:
                product = Product.objects.get(pk=product_id, is_active=True)
            except Product.DoesNotExist:
                errors.append(f"Product ID {product_id} not found")
                continue

            try:
                qty_str = quantities[i].strip() if i < len(quantities) else ''
                quantity = int(qty_str) if qty_str else 0
            except (ValueError, TypeError):
                errors.append(f"{product.name}: Invalid quantity")
                continue

            if quantity <= 0:
                continue

            movement_type = movement_types[i] if i < len(movement_types) else 'IN'
            adjustment_mode = adjustment_modes[i] if i < len(adjustment_modes) else 'PLUS'
            quantity_unit = quantity_units[i] if i < len(quantity_units) else 'NOS'
            # Always derive base cost from the selected product on the server.
            # This avoids mismatches from stale/tampered form values.
            cost_price = product.cost_price if product.cost_price is not None else Decimal('0.0')

            # Enforce manufacturer-specific cost maps so correct prices are always applied.
            normalized_name = normalize_sales_product_name(product.name).lower()
            if selected_manufacturer == 'Indian Kulfi':
                mapped_cost = IK_QUICK_ENTRY_COST_BY_NAME.get(normalized_name)
                if mapped_cost is not None:
                    cost_price = mapped_cost
            elif selected_manufacturer == 'Kulfi Corner':
                mapped_cost = KC_QUICK_ENTRY_COST_BY_NAME.get(normalized_name)
                if mapped_cost is not None:
                    cost_price = mapped_cost

            # Convert to actual units: pack = 6 units
            if quantity_unit == 'PACK':
                effective_quantity = quantity * 6
            else:
                effective_quantity = quantity

            movement_date_str = movement_dates[i] if i < len(movement_dates) else ''
            
            movement_value = Decimal(effective_quantity) * cost_price

            if movement_date_str:
                try:
                    movement_date = datetime.strptime(movement_date_str, '%Y-%m-%d').date()
                except ValueError:
                    movement_date = timezone.now().date()
            else:
                movement_date = timezone.now().date()

            # Stock operations
            if movement_type == 'OUT' and effective_quantity > product.current_stock:
                errors.append(f"{product.name}: Insufficient stock for outbound movement. Available {product.current_stock}, requested {effective_quantity}")
                continue

            inventory_quantity = effective_quantity
            if movement_type == 'IN':
                product.current_stock += effective_quantity
            elif movement_type == 'OUT':
                product.current_stock -= effective_quantity
            elif movement_type == 'ADJUSTMENT':
                stock_on_movement_date = calculate_stock_as_of_date(product, movement_date)
                if adjustment_mode == 'MINUS':
                    if effective_quantity > stock_on_movement_date:
                        errors.append(f"{product.name}: Cannot subtract {effective_quantity} on {movement_date}. Available stock on that date is {stock_on_movement_date}")
                        continue
                    adjusted_stock_on_date = stock_on_movement_date - effective_quantity
                    product.current_stock -= effective_quantity
                    adjustment_sign = '-'
                else:
                    adjusted_stock_on_date = stock_on_movement_date + effective_quantity
                    product.current_stock += effective_quantity
                    adjustment_sign = '+'

                # Keep ADJUSTMENT quantity as absolute stock snapshot after applying +/- change on movement_date.
                inventory_quantity = adjusted_stock_on_date
            else:
                errors.append(f"{product.name}: Unsupported movement type {movement_type}")
                continue

            # Persist inventory movement and product stock.
            # Stamp the manufacturer in notes so stock reports can identify it reliably.
            if selected_manufacturer:
                manufacturer_stamp = f'Manufacturer: {selected_manufacturer}'
            else:
                manufacturer_stamp = ''
            if movement_type == 'ADJUSTMENT':
                movement_notes = f'{manufacturer_stamp} | Adjustment mode: {adjustment_sign}{effective_quantity}' if manufacturer_stamp else f'Adjustment mode: {adjustment_sign}{effective_quantity}'
            else:
                movement_notes = manufacturer_stamp

            Inventory.objects.create(
                product=product,
                movement_type=movement_type,
                quantity=inventory_quantity,
                unit_cost=cost_price,
                movement_date=movement_date,
                reference_document=f'Quick entry {movement_date}',
                notes=movement_notes,
                created_by=request.user
            )
            product.current_stock = max(0, product.current_stock)
            product.save()
            recorded_items.append({
                'name': product.name,
                'sku': product.sku,
                'movement_type': movement_type,
                'quantity': quantity,
                'unit': quantity_unit,
                'effective_qty': effective_quantity,
                'adjustment_mode': adjustment_mode,
                'cost_price': str(cost_price),
                'total_value': str(movement_value),
                'movement_date': str(movement_date),
            })

            # Success-only summary totals for the post-submit green card.
            product_count += 1
            if quantity_unit == 'PACK':
                total_packs += quantity
            total_effective_qty += effective_quantity
            overall_total_value += movement_value
            created_count += 1

        if created_count > 0:
            messages.success(request, f'Successfully recorded {created_count} movement(s).')

        for error in errors:
            messages.warning(request, error)

        if created_count == 0 and not errors:
            messages.info(request, 'No inventory movements were recorded.')

        request.session['recorded_items'] = recorded_items
        request.session['recorded_summary'] = {
            'product_count': product_count,
            'total_packs': total_packs,
            'total_effective_qty': total_effective_qty,
            'overall_total_value': str(overall_total_value),
            'created_count': created_count,
        }
        return redirect('quick_inventory_entry')

    selected_movement_date_raw = request.GET.get('movement_date', '')
    selected_manufacturer = request.GET.get('manufacturer', '')
    selected_movement_date = None
    if selected_movement_date_raw:
        try:
            selected_movement_date = datetime.strptime(selected_movement_date_raw, '%Y-%m-%d').date()
        except ValueError:
            selected_movement_date = None

    # Build product groups: one entry per normalized flavor name, pairing IK and KC variants
    all_products = list(Product.objects.filter(is_active=True).order_by('sku'))
    stock_map = (
        get_stock_as_of_date_map(all_products, selected_movement_date)
        if selected_movement_date else
        {product.id: product.current_stock for product in all_products}
    )
    product_groups_map = {}
    for product in all_products:
        norm_name = normalize_sales_product_name(product.name)
        norm_key = norm_name.lower()
        if norm_key not in product_groups_map:
            product_groups_map[norm_key] = {
                'name': norm_name,
                'ik': None,
                'kc': None,
                'ik_stock': 0,
                'kc_stock': 0,
                'ik_cost_override': IK_QUICK_ENTRY_COST_BY_NAME.get(norm_key),
                'kc_cost_override': KC_QUICK_ENTRY_COST_BY_NAME.get(norm_key),
            }
        category = (product.category or '').strip().lower()
        if category == 'kulfi corner':
            product_groups_map[norm_key]['kc'] = product
            product_groups_map[norm_key]['kc_stock'] = max(0, stock_map.get(product.id, 0))
        else:
            product_groups_map[norm_key]['ik'] = product
            product_groups_map[norm_key]['ik_stock'] = max(0, stock_map.get(product.id, 0))

    for group in product_groups_map.values():
        # If only KC product exists for a flavor, reuse it for IK selection.
        # This supports shared inventory with manufacturer-specific costing.
        if group['ik'] is None and group['kc'] is not None:
            group['ik'] = group['kc']
            group['ik_stock'] = group['kc_stock']

    product_groups = sorted(
        product_groups_map.values(),
        key=lambda g: min(
            g['ik'].sku if g['ik'] else 'ZZZ999',
            g['kc'].sku if g['kc'] else 'ZZZ999',
        )
    )

    context = {
        'product_groups': product_groups,
        'today': timezone.now().date(),
        'selected_movement_date': selected_movement_date,
        'selected_manufacturer': selected_manufacturer,
        'recorded_items': request.session.pop('recorded_items', None),
        'recorded_summary': request.session.pop('recorded_summary', None),
    }
    return render(request, 'inventory/quick_inventory_entry.html', context)


@login_required
@permission_required('inventory.change_product', raise_exception=True)
def clear_stock(request, product_id):
    """Clear stock level for a product"""
    product = get_object_or_404(Product, pk=product_id, is_active=True)

    if request.method == 'POST':
        old_stock = product.current_stock
        if old_stock != 0:
            product.current_stock = 0
            product.save()

            Inventory.objects.create(
                product=product,
                movement_type='ADJUSTMENT',
                quantity=0,
                reference_document=f'Clear stock from {old_stock}',
                notes='Stock cleared via quick action',
                created_by=request.user
            )
            messages.success(request, f'Stock for {product.name} cleared to 0.')
        else:
            messages.info(request, f'Stock for {product.name} is already 0.')

        return redirect('inventory_list')

    context = {'product': product}
    return render(request, 'inventory/confirm_clear_stock.html', context)

@login_required
def inventory_history(request, product_id):
    """View inventory movement history for a product"""
    product = get_object_or_404(Product, pk=product_id)
    movements = Inventory.objects.filter(product=product)
    
    context = {
        'product': product,
        'movements': movements,
    }
    
    return render(request, 'inventory/inventory_history.html', context)

@login_required
def inventory_date_history(request):
    """View all inventory movements filtered by a date range."""
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')

    if not start_date_param or not end_date_param:
        # No complete date range submitted yet — show empty form
        return render(request, 'inventory/inventory_date_history.html', {
            'date_selected': False,
            'selected_start_date': '',
            'selected_end_date': '',
        })

    try:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
    except ValueError:
        start_date = timezone.now().date()

    try:
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
    except ValueError:
        end_date = timezone.now().date()

    if start_date > end_date:
        start_date, end_date = end_date, start_date
        messages.info(request, 'Start date and end date were swapped to apply a valid date range.')

    # Get all movements for the selected date range
    daily_movements = Inventory.objects.filter(
        movement_date__gte=start_date,
        movement_date__lte=end_date
    ).select_related('product', 'created_by').order_by('movement_date', 'product__sku', '-created_at')

    # Calculate stock summary
    total_in = daily_movements.filter(movement_type='IN').aggregate(
        total=Coalesce(Sum('quantity'), 0)
    )['total']
    total_out = daily_movements.filter(movement_type='OUT').aggregate(
        total=Coalesce(Sum('quantity'), 0)
    )['total']
    total_adjustment = daily_movements.filter(movement_type='ADJUSTMENT').aggregate(
        total=Coalesce(Sum('quantity'), 0)
    )['total']

    # Stock as of end date (Snapshot based on movements)
    stock_snapshot = []
    products = list(Product.objects.filter(is_active=True).order_by('sku'))

    # Apply custom kulfi display order
    _KULFI_SKU_ORDER = [
        'IK0001', 'IK0004', 'IK0005', 'IK0002', 'IK0003', 'IK0006',
        'IK0008', 'IK0011', 'IK0015', 'IK0012', 'IK0010', 'IK0007',
        'IK0009', 'IK0013', 'IK0014', 'IK0017', 'IK0018', 'IK0016',
    ]
    _sku_pos = {sku: i for i, sku in enumerate(_KULFI_SKU_ORDER)}
    products.sort(key=lambda p: _sku_pos.get(p.sku, len(_KULFI_SKU_ORDER)))

    stock_map = get_stock_as_of_date_map(products, end_date)

    for product in products:
        # Remove 'KC' from product name for display
        if 'KC' in product.name:
            product.name = product.name.replace('KC', '').strip()
        stock = stock_map.get(product.id, 0)
        stock_amount = Decimal(stock) * product.cost_price
        stock_snapshot.append({
            'product': product,
            'stock': stock,
            'stock_amount': stock_amount,
        })

    total_stock_as_of_date = sum(item['stock'] for item in stock_snapshot)
    total_stock_amount_as_of_date = sum(
        (item['stock_amount'] for item in stock_snapshot),
        Decimal('0.0')
    )

    context = {
        'date_selected': True,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'daily_movements': daily_movements,
        'total_movements': daily_movements.count(),
        'total_in': total_in,
        'total_out': total_out,
        'total_adjustment': total_adjustment,
        'stock_snapshot': stock_snapshot,
        'total_stock_as_of_date': total_stock_as_of_date,
        'total_stock_amount_as_of_date': total_stock_amount_as_of_date,
    }

    return render(request, 'inventory/inventory_date_history.html', context)


@login_required
def stock_order(request):
    """Stock order form for Indian Kulfi and Kulfi Corner manufacturers."""
    today = date.today().strftime('%Y-%m-%d')
    return render(request, 'inventory/stock_order.html', {'today': today})

# ==================== SALES MODULE ====================

def normalize_sales_product_name(name):
    """Normalize product name for manufacturer-agnostic sales grouping."""
    cleaned_name = re.sub(r'\s+', ' ', (name or '').strip())
    # Remove leading manufacturer tokens used in product names (e.g., 'IK Malai', 'KC Malai').
    cleaned_name = re.sub(r'^(IK|KC)\s*[-:]*\s*', '', cleaned_name, flags=re.IGNORECASE)
    cleaned_name = re.sub(r'\s*\((IK|KC)\)$', '', cleaned_name, flags=re.IGNORECASE)
    return cleaned_name.strip()


# Fixed product order required for View Sales by Date, Daily Report, and Weekly Report.
REPORT_PRODUCT_DISPLAY_ORDER = [
    'malai',
    'pista badam',
    'chocolate',
    'kesar badam',
    'kesar pista',
    'strawberry',
    'dry fruit',
    'blackcurrant',
    'litchi',
    'caramel coffee',
    'rose',
    'mango malai',
    'butter scotch',
    'coconut',
    'elaichi',
    'guava',
    'paan',
    'kajoor',
]
REPORT_PRODUCT_DISPLAY_INDEX = {
    name: index for index, name in enumerate(REPORT_PRODUCT_DISPLAY_ORDER)
}
REPORT_PRODUCT_DISPLAY_ALIASES = {
    'black current': 'blackcurrant',
    'black currant': 'blackcurrant',
    'blackcurrent': 'blackcurrant',
    'butterscotch': 'butter scotch',
    'elachi': 'elaichi',
    'kesar kajoor': 'kajoor',
}


def get_report_product_sort_key(product_name, sort_sku=''):
    """Return stable sort key using fixed report product order, then name/SKU fallback."""
    normalized_name = (product_name or '').strip().lower()
    canonical_name = REPORT_PRODUCT_DISPLAY_ALIASES.get(normalized_name, normalized_name)
    return (
        REPORT_PRODUCT_DISPLAY_INDEX.get(canonical_name, len(REPORT_PRODUCT_DISPLAY_INDEX)),
        canonical_name,
        sort_sku or '',
    )


# Fixed display order required for Daily Sales Sheet product rows.
DAILY_SALES_PRODUCT_DISPLAY_ORDER = [
    'malai',
    'pista badam',
    'chocolate',
    'kesar badam',
    'kesar pista',
    'strawberry',
    'dry fruit',
    'black currant',
    'litchi',
    'caramel coffee',
    'rose',
    'mango malai',
    'butterscotch',
    'coconut',
    'elaichi',
    'guava',
    'paan',
    'kesar kajoor',
]
DAILY_SALES_PRODUCT_DISPLAY_INDEX = {
    name: index for index, name in enumerate(DAILY_SALES_PRODUCT_DISPLAY_ORDER)
}
DAILY_SALES_PRODUCT_DISPLAY_ALIASES = {
    'black current': 'black currant',
    'blackcurrent': 'black currant',
    'straswberry': 'strawberry',
}
SALES_STOCK_TAKEN_PRODUCT_DISPLAY_ORDER = [
    'malai',
    'pista badam',
    'chocolate',
    'kesar badam',
    'kesar pista',
    'strawberry',
    'dry fruit',
    'black currant',
    'litchi',
    'caramel coffee',
    'rose',
    'mango malai',
    'butterscotch',
    'coconut',
    'elaichi',
    'guava',
    'paan',
    'kesar kajoor',
]
SALES_STOCK_TAKEN_PRODUCT_DISPLAY_INDEX = {
    name: index for index, name in enumerate(SALES_STOCK_TAKEN_PRODUCT_DISPLAY_ORDER)
}
SALES_STOCK_TAKEN_PRODUCT_DISPLAY_ALIASES = {
    'black current': 'black currant',
    'blackcurrant': 'black currant',
    'butter scotch': 'butterscotch',
    'elachi': 'elaichi',
    'kajoor': 'kesar kajoor',
    'straswberry': 'strawberry',
}
SALES_STOCK_TAKEN_PRODUCT_DISPLAY_LABELS = {
    'malai': 'MALAI',
    'pista badam': 'PISTA BADAM',
    'chocolate': 'CHOCOLATE',
    'kesar badam': 'KESAR BADAM',
    'kesar pista': 'KESAR PISTA',
    'strawberry': 'STRAWBERRY',
    'dry fruit': 'DRY FRUIT',
    'black currant': 'BLACK CURRANT',
    'litchi': 'LITCHI',
    'caramel coffee': 'CARAMEL COFFEE',
    'rose': 'ROSE',
    'mango malai': 'MANGO MALAI',
    'butterscotch': 'BUTTERSCOTCH',
    'coconut': 'COCONUT',
    'elaichi': 'ELAICHI',
    'guava': 'GUAVA',
    'paan': 'PAAN',
    'kesar kajoor': 'KESAR KAJOOR',
}


def get_sales_stock_taken_product_name_key(product_name):
    normalized_name = (product_name or '').strip().lower()
    return SALES_STOCK_TAKEN_PRODUCT_DISPLAY_ALIASES.get(normalized_name, normalized_name)


def get_sales_stock_taken_product_sort_key(product_name, sort_sku=''):
    canonical_name = get_sales_stock_taken_product_name_key(product_name)
    return (
        SALES_STOCK_TAKEN_PRODUCT_DISPLAY_INDEX.get(canonical_name, len(SALES_STOCK_TAKEN_PRODUCT_DISPLAY_INDEX)),
        canonical_name,
        sort_sku or '',
    )


def get_sales_stock_taken_product_display_name(product_name):
    canonical_name = get_sales_stock_taken_product_name_key(product_name)
    return SALES_STOCK_TAKEN_PRODUCT_DISPLAY_LABELS.get(canonical_name, canonical_name.upper())


def group_active_products_by_name():
    """Return active products grouped by normalized product name key."""
    grouped = defaultdict(list)
    for product in Product.objects.filter(is_active=True):
        key = normalize_sales_product_name(product.name).lower()
        grouped[key].append(product)
    return grouped


def build_sales_groups(sales_qs, include_date=False):
    """Aggregate sales rows by normalized product name (and date if requested)."""
    grouped = {}
    for sale in sales_qs:
        display_name = normalize_sales_product_name(sale.product.name)
        group_key = (sale.sale_date, display_name) if include_date else (display_name,)

        if group_key not in grouped:
            grouped[group_key] = {
                'sale_date': sale.sale_date,
                'product_name': display_name,
                'quantity': 0,
                'total_price': Decimal('0.0'),
                'recorded_by': set(),
                'notes': [],
                'sort_sku': sale.product.sku,
            }

        grouped[group_key]['quantity'] += sale.quantity
        grouped[group_key]['total_price'] += sale.total_price
        grouped[group_key]['recorded_by'].add(
            sale.recorded_by.get_full_name() or sale.recorded_by.username
        )
        if sale.notes:
            grouped[group_key]['notes'].append(sale.notes)

        # Keep lowest SKU in each group to support SKU-based ordering.
        current_sort_sku = grouped[group_key].get('sort_sku')
        if not current_sort_sku or sale.product.sku < current_sort_sku:
            grouped[group_key]['sort_sku'] = sale.product.sku

    grouped_rows = []
    for _, row in grouped.items():
        row['unit_price'] = (
            row['total_price'] / row['quantity'] if row['quantity'] else Decimal('0.0')
        )
        row['recorded_by'] = ', '.join(sorted(row['recorded_by'])) if row['recorded_by'] else '-'
        row['notes'] = ' | '.join(row['notes']) if row['notes'] else '-'
        grouped_rows.append(row)

    if include_date:
        grouped_rows.sort(
            key=lambda item: (
                item['sale_date'],
                *get_report_product_sort_key(item['product_name'], item.get('sort_sku', '')),
            )
        )
    else:
        grouped_rows.sort(
            key=lambda item: get_report_product_sort_key(item['product_name'], item.get('sort_sku', ''))
        )
    return grouped_rows


def _recalculate_current_stock_for_products(product_ids):
    """Rebuild current stock from inventory and sales history for the affected products."""
    unique_product_ids = {product_id for product_id in product_ids if product_id}
    if not unique_product_ids:
        return

    today = timezone.now().date()
    for product in Product.objects.filter(id__in=unique_product_ids):
        product.current_stock = max(0, calculate_stock_as_of_date(product, today))
        product.save(update_fields=['current_stock'])


def build_grouped_products_for_sales_date(selected_sales_date):
    """Build grouped product rows with combined stock and average unit price for a date."""
    grouped_products = defaultdict(list)
    for _, products_in_group in group_active_products_by_name().items():
        if products_in_group:
            group_name = normalize_sales_product_name(products_in_group[0].name)
            grouped_products[group_name] = products_in_group

    grouped_products_for_form = []
    grouped_items = sorted(
        grouped_products.items(),
        key=lambda item: get_sales_stock_taken_product_sort_key(
            item[0],
            min((product.sku for product in item[1]), default='ZZZ999'),
        )
    )
    for group_name, products_in_group in grouped_items:
        stock_map = get_stock_as_of_date_map(products_in_group, selected_sales_date)
        total_stock = max(0, sum(stock_map.get(item.id, 0) for item in products_in_group))
        avg_price = (
            sum(item.selling_price for item in products_in_group) / len(products_in_group)
            if products_in_group else Decimal('0.0')
        )
        grouped_products_for_form.append({
            'key': get_sales_stock_taken_product_name_key(group_name),
            'name': get_sales_stock_taken_product_display_name(group_name),
            'stock': total_stock,
            'avg_price': avg_price,
        })

    return grouped_products_for_form


@login_required
def sales_stock_taken_entry(request):
    """Capture stock taken by salespeople without reducing inventory stock."""
    if not request.user.is_staff:
        messages.error(request, 'Use Daily Sales Sheet for sales stock and sales updates.')
        return redirect('quick_sales_entry')

    selected_sales_date_raw = request.POST.get('sales_date') or request.GET.get('sales_date') or request.GET.get('date')
    if selected_sales_date_raw:
        try:
            selected_sales_date = datetime.strptime(selected_sales_date_raw, '%Y-%m-%d').date()
        except ValueError:
            selected_sales_date = timezone.now().date()
    else:
        selected_sales_date = timezone.now().date()

    grouped_products_for_form = build_grouped_products_for_sales_date(selected_sales_date)
    products_by_key = {item['key']: item for item in grouped_products_for_form}

    target_user = request.user
    selected_salesperson_id = ''
    salespeople = User.objects.filter(is_staff=False, is_active=True).order_by('first_name', 'username')
    if request.user.is_staff:
        selected_salesperson_id = request.POST.get('salesperson') or request.GET.get('salesperson', '')
        if selected_salesperson_id:
            try:
                target_user = User.objects.get(pk=int(selected_salesperson_id), is_active=True)
            except (User.DoesNotExist, ValueError, TypeError):
                target_user = request.user
                selected_salesperson_id = ''
                messages.warning(request, 'Selected salesperson not found. Showing your own records.')

    if request.method == 'POST':
        product_keys = request.POST.getlist('product_key[]')
        stock_taken_counts = request.POST.getlist('stock_taken_count[]')

        saved_count = 0
        removed_count = 0
        errors = []

        existing_entries = {
            get_sales_stock_taken_product_name_key(item.product_key or item.product_name): item
            for item in SalesStockTaken.objects.filter(
                salesperson=target_user,
                sales_date=selected_sales_date,
            )
        }

        for i, product_key in enumerate(product_keys):
            if not product_key:
                continue

            product_info = products_by_key.get(product_key)
            if not product_info:
                errors.append(f'Row {i + 1}: Product group not found.')
                continue

            raw_count = stock_taken_counts[i].strip() if i < len(stock_taken_counts) else '0'
            if not raw_count:
                raw_count = '0'

            try:
                stock_taken_count = int(raw_count)
            except (TypeError, ValueError):
                errors.append(f"{product_info['name']}: Invalid stock taken count")
                continue

            if stock_taken_count < 0:
                errors.append(f"{product_info['name']}: Stock taken count cannot be negative")
                continue

            if stock_taken_count > 0:
                existing_entry = existing_entries.get(product_key)
                if existing_entry:
                    existing_entry.product_key = product_key
                    existing_entry.product_name = product_info['name']
                    existing_entry.avg_unit_price = product_info['avg_price']
                    existing_entry.combined_stock = product_info['stock']
                    existing_entry.stock_taken_count = stock_taken_count
                    existing_entry.save()
                else:
                    SalesStockTaken.objects.update_or_create(
                        salesperson=target_user,
                        sales_date=selected_sales_date,
                        product_key=product_key,
                        defaults={
                            'product_name': product_info['name'],
                            'avg_unit_price': product_info['avg_price'],
                            'combined_stock': product_info['stock'],
                            'stock_taken_count': stock_taken_count,
                        },
                    )
                saved_count += 1
            else:
                existing_entry = existing_entries.get(product_key)
                if existing_entry:
                    existing_entry.delete()
                    removed_count += 1

        if saved_count > 0:
            messages.success(request, f'Saved stock taken records for {saved_count} product(s).')
        if removed_count > 0:
            messages.info(request, f'Removed {removed_count} empty stock taken record(s).')
        for error in errors:
            messages.warning(request, error)

        if saved_count == 0 and removed_count == 0 and not errors:
            messages.info(request, 'No stock taken records were updated.')

        redirect_url = f"{reverse('sales_stock_taken_entry')}?sales_date={selected_sales_date.isoformat()}"
        if request.user.is_staff and target_user.id != request.user.id:
            redirect_url = f"{redirect_url}&salesperson={target_user.id}"
        return redirect(redirect_url)

    existing_entries = {
        get_sales_stock_taken_product_name_key(item.product_key or item.product_name): item
        for item in SalesStockTaken.objects.filter(
            salesperson=target_user,
            sales_date=selected_sales_date,
        )
    }

    admin_day_entries = None
    if request.user.is_staff:
        admin_day_entries = sorted(
            SalesStockTaken.objects.filter(
                sales_date=selected_sales_date,
            ).select_related('salesperson'),
            key=lambda entry: (
                entry.salesperson.username,
                *get_sales_stock_taken_product_sort_key(entry.product_name),
            ),
        )
        for entry in admin_day_entries:
            entry.display_product_name = get_sales_stock_taken_product_display_name(entry.product_name)

    total_taken_count = 0
    total_estimated_value = Decimal('0.0')
    for product in grouped_products_for_form:
        existing = existing_entries.get(product['key'])
        if existing:
            stock_taken_count = existing.stock_taken_count
        else:
            stock_taken_count = 0
        product['stock_taken_count'] = stock_taken_count
        product['estimated_total'] = product['avg_price'] * stock_taken_count
        total_taken_count += stock_taken_count
        total_estimated_value += product['estimated_total']

    total_combined_stock = sum(product['stock'] for product in grouped_products_for_form)
    context = {
        'products': grouped_products_for_form,
        'selected_sales_date': selected_sales_date,
        'today': timezone.now().date(),
        'total_taken_count': total_taken_count,
        'total_estimated_value': total_estimated_value,
        'total_combined_stock': total_combined_stock,
        'target_user': target_user,
        'salespeople': salespeople,
        'selected_salesperson_id': selected_salesperson_id,
        'admin_day_entries': admin_day_entries,
    }
    return render(request, 'inventory/sales_stock_taken_entry.html', context)

@login_required
def quick_sales_entry(request):
    """Daily sales sheet by product name (manufacturer-agnostic)."""
    def get_sales_groups():
        """Return grouped active products across all categories."""
        grouped = defaultdict(list)
        for product in Product.objects.filter(is_active=True):
            key = get_sales_stock_taken_product_name_key(normalize_sales_product_name(product.name))
            grouped[key].append(product)
        return grouped


    if request.method == 'POST':
        action = request.POST.get('action', 'record_sales')

        if request.user.is_staff and action == 'save_stock_taken':
            messages.error(request, 'Stock taken entries are only available for sales users.')
            return redirect('quick_sales_entry')


        # Handle multiple sales entries by normalized product name.
        product_keys = request.POST.getlist('product_key[]')
        product_labels = request.POST.getlist('product_label[]')
        quantities = request.POST.getlist('quantity[]')
        stock_taken_counts = request.POST.getlist('stock_taken_count[]')
        sale_dates = request.POST.getlist('sale_date[]')
        notes_list = request.POST.getlist('notes[]')

        # --- Save Sales Count Draft ---
        if action == 'save_sales_count' and not request.user.is_staff:
            selected_sales_date = None
            if sale_dates and sale_dates[0]:
                try:
                    selected_sales_date = datetime.strptime(sale_dates[0], '%Y-%m-%d').date()
                except Exception:
                    selected_sales_date = timezone.now().date()
            else:
                selected_sales_date = timezone.now().date()

            saved_count = 0
            for i, product_key in enumerate(product_keys):
                if not product_key:
                    continue
                try:
                    quantity = int(quantities[i]) if i < len(quantities) else 0
                except Exception:
                    quantity = 0
                # Save or update draft
                obj, created = SalesCountDraft.objects.update_or_create(
                    salesperson=request.user,
                    sales_date=selected_sales_date,
                    product_key=product_key,
                    defaults={
                        'sales_count': quantity,
                    },
                )
                saved_count += 1
            messages.success(request, f'Saved sales count draft for {saved_count} product(s).')
            return redirect(f"{reverse('quick_sales_entry')}?sales_date={selected_sales_date}")

        # If a sales user records sales without entering any sales count,
        # assume all previously saved stock taken for the selected date is sold.
        if action == 'record_sales' and not request.user.is_staff:
            entered_positive_quantity = False
            for i in range(len(product_keys)):
                try:
                    entered_qty = int((quantities[i] if i < len(quantities) else '0') or 0)
                except (ValueError, TypeError):
                    entered_qty = 0
                if entered_qty > 0:
                    entered_positive_quantity = True
                    break

            if not entered_positive_quantity:
                fallback_sale_date = timezone.now().date()
                for sale_date_str in sale_dates:
                    if sale_date_str:
                        try:
                            fallback_sale_date = datetime.strptime(sale_date_str, '%Y-%m-%d').date()
                            break
                        except ValueError:
                            continue

                stock_taken_lookup = {
                    get_sales_stock_taken_product_name_key(item.product_key or item.product_name): item.stock_taken_count
                    for item in SalesStockTaken.objects.filter(
                        salesperson=request.user,
                        sales_date=fallback_sale_date,
                    )
                }

                if stock_taken_lookup:
                    redirect_url = (
                        f"{reverse('quick_sales_entry')}"
                        f"?sales_date={fallback_sale_date.isoformat()}"
                        f"&prefill_from_stock_taken=1"
                    )
                    messages.info(
                        request,
                        'Sales Count was left empty, so saved Stock Taken counts were copied to Sales Count. '
                        'Review and click Record Sales again to confirm.'
                    )
                    return redirect(redirect_url)

        grouped_products = get_sales_groups()
        touched_product_ids = set()

        sales_created = 0
        errors = []
        total_items_sold = 0
        total_selling_price = Decimal('0.0')
        recorded_sale_dates = set()
        stock_taken_saved = 0
        stock_taken_removed = 0

        # Save stock taken values only when explicit action is requested.
        if action == 'save_stock_taken':
            for i, product_key in enumerate(product_keys):
                if not product_key:
                    continue

                try:
                    sale_date_str = sale_dates[i] if i < len(sale_dates) else ''
                    if sale_date_str:
                        try:
                            sale_date = datetime.strptime(sale_date_str, '%Y-%m-%d').date()
                        except ValueError:
                            sale_date = timezone.now().date()
                    else:
                        sale_date = timezone.now().date()

                    candidates = grouped_products.get(product_key, [])
                    label = product_labels[i] if i < len(product_labels) else product_key

                    if not candidates:
                        errors.append(f"{label}: Product not found in active inventory")
                        continue

                    total_available_stock = sum(max(0, item.current_stock) for item in candidates)

                    raw_stock_taken = stock_taken_counts[i].strip() if i < len(stock_taken_counts) else '0'
                    if not raw_stock_taken:
                        raw_stock_taken = '0'

                    try:
                        stock_taken_count = int(raw_stock_taken)
                        if stock_taken_count < 0:
                            errors.append(f"{label}: Stock taken count cannot be negative")
                            stock_taken_count = 0
                    except (ValueError, TypeError):
                        errors.append(f"{label}: Invalid stock taken count")
                        stock_taken_count = 0

                    avg_price = (
                        sum(item.selling_price for item in candidates) / len(candidates)
                        if candidates else Decimal('0.0')
                    )

                    if stock_taken_count > 0:
                        SalesStockTaken.objects.update_or_create(
                            salesperson=request.user,
                            sales_date=sale_date,
                            product_key=product_key,
                            defaults={
                                'product_name': label,
                                'avg_unit_price': avg_price,
                                'combined_stock': total_available_stock,
                                'stock_taken_count': stock_taken_count,
                            },
                        )
                        stock_taken_saved += 1
                    else:
                        deleted_count, _ = SalesStockTaken.objects.filter(
                            salesperson=request.user,
                            sales_date=sale_date,
                            product_key=product_key,
                        ).delete()
                        if deleted_count:
                            stock_taken_removed += 1
                except Exception as e:
                    errors.append(f"Error processing row {i+1}: {str(e)}")

            if stock_taken_saved > 0:
                messages.success(request, f'Saved stock taken records for {stock_taken_saved} product(s).')
            if stock_taken_removed > 0:
                messages.info(request, f'Removed {stock_taken_removed} empty stock taken record(s).')
            if errors:
                for error in errors:
                    messages.warning(request, f'⚠ {error}')
            if stock_taken_saved == 0 and stock_taken_removed == 0 and not errors:
                messages.info(request, 'No stock taken records were updated.')

            return redirect('quick_sales_entry')


        for i, product_key in enumerate(product_keys):
            if not product_key:
                continue

            try:
                try:
                    quantity = int(quantities[i]) if i < len(quantities) else 0
                except (ValueError, IndexError):
                    label = product_labels[i] if i < len(product_labels) else 'Product'
                    errors.append(f"{label}: Invalid quantity value")
                    continue

                sale_date_str = sale_dates[i] if i < len(sale_dates) else ''
                if sale_date_str:
                    try:
                        sale_date = datetime.strptime(sale_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        sale_date = timezone.now().date()
                else:
                    sale_date = timezone.now().date()

                notes = notes_list[i] if i < len(notes_list) else ''
                candidates = grouped_products.get(product_key, [])
                label = product_labels[i] if i < len(product_labels) else product_key

                if not candidates:
                    errors.append(f"{label}: Product not found in active inventory")
                    continue

                total_available_stock = sum(max(0, item.current_stock) for item in candidates)

                if quantity <= 0:
                    continue

                if quantity > total_available_stock:
                    errors.append(
                        f"{label}: Insufficient stock. Available: {total_available_stock}, Requested: {quantity}"
                    )
                    continue

                remaining_quantity = quantity
                candidates_sorted = sorted(
                    candidates,
                    key=lambda item: max(0, item.current_stock),
                    reverse=True,
                )
                for product in candidates_sorted:
                    if remaining_quantity <= 0:
                        break

                    available_for_date = max(0, product.current_stock)
                    allocated_quantity = min(remaining_quantity, available_for_date)
                    if allocated_quantity <= 0:
                        continue

                    Sales.objects.create(
                        product=product,
                        quantity=allocated_quantity,
                        unit_price=product.selling_price,
                        sale_date=sale_date,
                        recorded_by=request.user,
                        notes=notes,
                    )

                    product.current_stock -= allocated_quantity
                    product.current_stock = max(0, product.current_stock)
                    product.save()

                    total_items_sold += allocated_quantity
                    total_selling_price += Decimal(allocated_quantity) * product.selling_price
                    sales_created += 1
                    remaining_quantity -= allocated_quantity
                    touched_product_ids.add(product.id)

                recorded_sale_dates.add(sale_date)
            except Exception as e:
                errors.append(f"Error processing row {i+1}: {str(e)}")

        # --- Clear sales count drafts for this user/date after recording sales ---
        if not request.user.is_staff:
            # Use the first sale_date in sale_dates or today
            clear_date = None
            if sale_dates and sale_dates[0]:
                try:
                    clear_date = datetime.strptime(sale_dates[0], '%Y-%m-%d').date()
                except Exception:
                    clear_date = timezone.now().date()
            else:
                clear_date = timezone.now().date()
            SalesCountDraft.objects.filter(salesperson=request.user, sales_date=clear_date).delete()

        # Rebuild live stock from historical records after backdated inserts.
        if touched_product_ids:
            _recalculate_current_stock_for_products(touched_product_ids)

        # Show results to user
        if sales_created > 0:
            if len(recorded_sale_dates) == 1:
                only_date = next(iter(recorded_sale_dates))
                date_text = only_date.strftime('%Y-%m-%d')
            else:
                sorted_dates = sorted(recorded_sale_dates)
                date_text = f"{sorted_dates[0].strftime('%Y-%m-%d')} to {sorted_dates[-1].strftime('%Y-%m-%d')}"

            messages.success(request, (
                f'Successfully recorded {sales_created} sales, '
                f'{total_items_sold} items sold, '
                f'sales date: {date_text}, '
                f'total selling price ₹{total_selling_price:.2f}'
            ))
        
        if errors:
            for error in errors:
                messages.warning(request, f'⚠ {error}')
        
        if sales_created == 0 and not errors:
            messages.info(request, 'No sales were recorded.')

        return redirect('quick_sales_entry')


    # GET request - show form with stock as-of selected sales date
    selected_sales_date_raw = request.GET.get('sales_date') or request.GET.get('date')
    if selected_sales_date_raw:
        try:
            selected_sales_date = datetime.strptime(selected_sales_date_raw, '%Y-%m-%d').date()
        except ValueError:
            selected_sales_date = timezone.now().date()
    else:
        selected_sales_date = timezone.now().date()

    grouped_products_for_form = build_grouped_products_for_sales_date(selected_sales_date)
    prefill_from_stock_taken = request.GET.get('prefill_from_stock_taken') == '1'

    # --- Prefill sales count from draft if exists ---
    sales_count_drafts = {}
    if not request.user.is_staff:
        for draft in SalesCountDraft.objects.filter(salesperson=request.user, sales_date=selected_sales_date):
            sales_count_drafts[draft.product_key] = draft.sales_count

    # Daily Sales Sheet should always reflect live stock, even for backdated sale dates.
    current_grouped_products = get_sales_groups()
    # Use historical stock as of selected_sales_date for combined stock
    for product in grouped_products_for_form:
        candidates = current_grouped_products.get(product['key'], [])
        stock_map = get_stock_as_of_date_map(candidates, selected_sales_date)
        product['stock'] = sum(stock_map.get(item.id, 0) for item in candidates)

    stock_taken_map = {
        get_sales_stock_taken_product_name_key(item.product_key or item.product_name): item.stock_taken_count
        for item in SalesStockTaken.objects.filter(
            salesperson=request.user,
            sales_date=selected_sales_date,
        )
    }

    for product in grouped_products_for_form:
        product['stock_taken_count'] = stock_taken_map.get(product['key'], 0)
        # Prefill logic: draft > stock_taken (if prefill_from_stock_taken) > 0
        if not request.user.is_staff:
            if product['key'] in sales_count_drafts:
                product['prefill_quantity'] = sales_count_drafts[product['key']]
            elif prefill_from_stock_taken:
                product['prefill_quantity'] = stock_taken_map.get(product['key'], 0)
            else:
                product['prefill_quantity'] = 0
        else:
            product['prefill_quantity'] = 0

    total_stock_taken_for_date = sum(stock_taken_map.values())

    total_combined_stock = sum(product['stock'] for product in grouped_products_for_form)
    context = {
        'products': grouped_products_for_form,
        'today': timezone.now().date(),
        'selected_sales_date': selected_sales_date,
        'total_stock_taken_for_date': total_stock_taken_for_date,
        'total_combined_stock': total_combined_stock,
        'prefill_from_stock_taken': prefill_from_stock_taken,
    }
    return render(request, 'inventory/quick_sales_entry.html', context)

@login_required
def view_sales(request):
    """View sales for a specific date"""
    if not request.user.is_staff:
        messages.error(request, 'You can access only Daily Sales Sheet with this account.')
        return redirect('quick_sales_entry')

    date_submitted = 'date' in request.GET
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    selected_salesperson_id = request.GET.get('salesperson', '').strip()
    restored_units_raw = (request.GET.get('restored_units') or '').strip()
    restored_product = (request.GET.get('restored_product') or '').strip()

    try:
        restored_units = int(restored_units_raw) if restored_units_raw else 0
    except ValueError:
        restored_units = 0

    if isinstance(selected_date, str):
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()

    previous_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)

    salespeople = User.objects.filter(
        id__in=Sales.objects.filter(
            sale_date=selected_date,
            recorded_by__isnull=False,
        ).values_list('recorded_by_id', flat=True).distinct()
    ).order_by('first_name', 'username')

    salesperson_filter = None
    if selected_salesperson_id:
        try:
            salesperson_filter = salespeople.get(pk=int(selected_salesperson_id))
        except (ValueError, User.DoesNotExist):
            selected_salesperson_id = ''
            if date_submitted:
                messages.warning(request, 'Selected salesperson was not found for this date.')

    if date_submitted:
        # Get sales for the selected date and group by product name.
        daily_sales_qs = Sales.objects.filter(sale_date=selected_date)
        if salesperson_filter:
            daily_sales_qs = daily_sales_qs.filter(recorded_by=salesperson_filter)
        daily_sales_qs = daily_sales_qs.select_related('product', 'recorded_by').order_by('product__sku')
        grouped_daily_sales = build_sales_groups(daily_sales_qs)

        # Calculate totals
        total_sales_count = len(grouped_daily_sales)
        total_revenue = daily_sales_qs.aggregate(
            total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
        )['total']
        total_quantity = daily_sales_qs.aggregate(
            total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
        )['total']
    else:
        grouped_daily_sales = None
        total_sales_count = 0
        total_revenue = 0
        total_quantity = 0

    context = {
        'selected_date': selected_date,
        'previous_date': previous_date,
        'next_date': next_date,
        'selected_salesperson_id': selected_salesperson_id,
        'selected_salesperson': salesperson_filter,
        'salespeople': salespeople,
        'date_submitted': date_submitted,
        'daily_sales': grouped_daily_sales,
        'total_sales_count': total_sales_count,
        'total_revenue': total_revenue,
        'total_quantity': total_quantity,
        'restored_units': max(0, restored_units),
        'restored_product': restored_product,
    }

    return render(request, 'inventory/view_sales.html', context)

@login_required
def edit_sale(request, sale_id):
    """Edit a sales record (admin only)"""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to edit sales.')
        return redirect('view_sales')
    
    sale = get_object_or_404(Sales, pk=sale_id)
    
    if request.method == 'POST':
        quantity = request.POST.get('quantity', sale.quantity)
        sale_date = request.POST.get('sale_date', sale.sale_date)
        notes = request.POST.get('notes', sale.notes)
        
        try:
            quantity = int(quantity)
            
            if quantity <= 0:
                messages.error(request, 'Quantity must be greater than 0.')
                return render(request, 'inventory/edit_sale.html', {'sale': sale})
            
            # Calculate stock difference
            old_quantity = sale.quantity
            quantity_diff = quantity - old_quantity
            product = sale.product
            
            # Check if sufficient stock available
            if quantity_diff > 0 and quantity_diff > product.current_stock:
                messages.error(request, f'Insufficient stock. Available: {product.current_stock}')
                return render(request, 'inventory/edit_sale.html', {'sale': sale})
            
            # Update sale
            sale.quantity = quantity
            sale.sale_date = sale_date
            sale.notes = notes
            sale.total_price = quantity * sale.unit_price
            sale.save()
            
            # Update stock
            product.current_stock -= quantity_diff
            product.current_stock = max(0, product.current_stock)
            product.save()
            
            messages.success(request, 'Sale updated successfully.')
            return redirect('view_sales', )
        
        except ValueError:
            messages.error(request, 'Invalid quantity value.')
            return render(request, 'inventory/edit_sale.html', {'sale': sale})
        except Exception as e:
            messages.error(request, f'Error updating sale: {str(e)}')
            return render(request, 'inventory/edit_sale.html', {'sale': sale})
    
    context = {'sale': sale}
    return render(request, 'inventory/edit_sale.html', context)

@login_required
def delete_sale(request, sale_id):
    """Delete a sales record (admin only) with stock restoration"""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to delete sales.')
        return redirect('view_sales')
    
    sale = get_object_or_404(Sales, pk=sale_id)
    
    if request.method == 'POST':
        product_id = sale.product_id
        product_name = sale.product.name
        deleted_quantity = sale.quantity

        with transaction.atomic():
            sale.delete()
            _recalculate_current_stock_for_products([product_id])

        messages.success(
            request,
            f'Sale deleted successfully. Removed {deleted_quantity} units from {product_name} sales history and recalculated inventory.'
        )
        return redirect('view_sales', )
    
    context = {'sale': sale}
    return render(request, 'inventory/confirm_delete_sale.html', context)


@login_required
@require_POST
def delete_grouped_sale(request):
    """Delete grouped sales row from View Sales and rebuild current stock."""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to delete sales.')
        return redirect('view_sales')

    selected_date_raw = (request.POST.get('selected_date') or '').strip()
    product_name_raw = (request.POST.get('product_name') or '').strip()
    selected_salesperson_id = (request.POST.get('selected_salesperson_id') or '').strip()

    if not selected_date_raw or not product_name_raw:
        messages.error(request, 'Missing sale date or product name for deletion.')
        return redirect('view_sales')

    try:
        selected_date = datetime.strptime(selected_date_raw, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid sale date for deletion.')
        return redirect('view_sales')

    sales_qs = Sales.objects.filter(sale_date=selected_date).select_related('product')

    if selected_salesperson_id:
        try:
            sales_qs = sales_qs.filter(recorded_by_id=int(selected_salesperson_id))
        except ValueError:
            messages.error(request, 'Invalid salesperson for deletion.')
            return redirect('view_sales')

    target_key = normalize_sales_product_name(product_name_raw).lower()
    matching_sales = [
        sale for sale in sales_qs
        if normalize_sales_product_name(sale.product.name).lower() == target_key
    ]

    if not matching_sales:
        messages.warning(request, 'No matching sales records were found to delete.')
    else:
        restored_quantity = 0
        touched_product_ids = []
        sale_ids = []
        for sale in matching_sales:
            restored_quantity += sale.quantity
            touched_product_ids.append(sale.product_id)
            sale_ids.append(sale.id)

        with transaction.atomic():
            Sales.objects.filter(id__in=sale_ids).delete()
            _recalculate_current_stock_for_products(touched_product_ids)

        messages.success(
            request,
            f'Deleted sales for {product_name_raw}. Removed {restored_quantity} units from sales history and recalculated inventory.'
        )

    query_params = {'date': selected_date.isoformat()}
    if selected_salesperson_id:
        query_params['salesperson'] = selected_salesperson_id
    if matching_sales:
        query_params['restored_units'] = restored_quantity
        query_params['restored_product'] = product_name_raw

    return redirect(f"{reverse('view_sales')}?{urlencode(query_params)}")


@login_required
@require_POST
def delete_sales_for_date(request):
    """Delete all sales recorded on a selected date and rebuild current stock."""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to delete sales.')
        return redirect('view_sales')

    selected_date_raw = (request.POST.get('selected_date') or '').strip()
    selected_salesperson_id = (request.POST.get('selected_salesperson_id') or '').strip()

    if not selected_date_raw:
        messages.error(request, 'Missing sale date for deletion.')
        return redirect('view_sales')

    try:
        selected_date = datetime.strptime(selected_date_raw, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid sale date for deletion.')
        return redirect('view_sales')

    sales_qs = Sales.objects.filter(sale_date=selected_date)
    sale_ids = list(sales_qs.values_list('id', flat=True))

    if not sale_ids:
        messages.warning(request, 'No sales records were found for the selected date.')
    else:
        touched_product_ids = list(sales_qs.values_list('product_id', flat=True))
        deleted_entries = len(sale_ids)
        deleted_units = sales_qs.aggregate(
            total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
        )['total']

        with transaction.atomic():
            Sales.objects.filter(id__in=sale_ids).delete()
            _recalculate_current_stock_for_products(touched_product_ids)

        messages.success(
            request,
            f'Deleted all {deleted_entries} sales entries on {selected_date.isoformat()} totaling {deleted_units} units. Inventory recalculated.'
        )

    query_params = {'date': selected_date.isoformat()}
    if selected_salesperson_id:
        query_params['salesperson'] = selected_salesperson_id

    return redirect(f"{reverse('view_sales')}?{urlencode(query_params)}")

@login_required
def sales_history(request):
    """View complete sales history with filtering on a single page."""
    if not request.user.is_staff:
        messages.error(request, 'You can access only Daily Sales Sheet with this account.')
        return redirect('quick_sales_entry')

    clear_filters = request.GET.get('clear') == '1'

    if clear_filters:
        context = {
            'sales': Sales.objects.none(),
            'start_date': '',
            'end_date': '',
            'total_sales_value': Decimal('0.0'),
            'total_sales_count': 0,
            'total_items_sold': 0,
        }
        return render(request, 'inventory/sales_history.html', context)

    sales_qs = Sales.objects.all().select_related('product', 'recorded_by')

    # Filter by date range if provided
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # If no date range provided at all, show empty state
    if not start_date and not end_date:
        context = {
            'sales': [],
            'start_date': '',
            'end_date': '',
            'total_sales_value': Decimal('0.0'),
            'total_sales_count': 0,
            'total_items_sold': 0,
            'no_filter': True,
        }
        return render(request, 'inventory/sales_history.html', context)

    if start_date:
        sales_qs = sales_qs.filter(sale_date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(sale_date__lte=end_date)

    sales = build_sales_groups(sales_qs, include_date=True)

    # Calculate total sales value for filtered results
    total_sales_value = sales_qs.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']
    total_sales_count = len(sales)
    total_items_sold = sales_qs.aggregate(
        total=Coalesce(Sum('quantity'), 0)
    )['total']

    context = {
        'sales': sales,
        'start_date': start_date,
        'end_date': end_date,
        'total_sales_value': total_sales_value,
        'total_sales_count': total_sales_count,
        'total_items_sold': total_items_sold,
    }

    return render(request, 'inventory/sales_history.html', context)

@login_required
def get_product_price(request):
    """AJAX endpoint to get product selling price"""
    product_id = request.GET.get('product_id')
    try:
        product = Product.objects.get(pk=product_id)
        return JsonResponse({
            'success': True,
            'price': float(product.selling_price),
            'stock': max(0, product.current_stock)
        })
    except Product.DoesNotExist:
        return JsonResponse({'success': False})


@login_required
def get_next_sku(request):
    """AJAX endpoint to get next SKU based on selected category."""
    category = request.GET.get('category', '')
    if not category:
        return JsonResponse({'success': False, 'message': 'Category is required'})

    next_sku = ProductForm.generate_next_sku(category)
    prefix = ProductForm.get_category_prefix(category)
    return JsonResponse({'success': True, 'sku': next_sku, 'prefix': prefix})


# ==================== OPERATIONS MODULE ====================

def _can_manage_operation_expense(user, expense):
    return user.is_staff or expense.created_by_id == user.id


def _can_manage_operation_income(user, income):
    return user.is_staff or income.created_by_id == user.id

@login_required
def quick_operations_entry(request):
    """Quick entry for dated operational expenses."""
    selected_date_raw = request.GET.get('date')
    if selected_date_raw:
        try:
            selected_date = datetime.fromisoformat(selected_date_raw).date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()

    edit_expense = None
    edit_id_raw = request.GET.get('edit')
    if edit_id_raw:
        try:
            edit_expense = OperationsExpense.objects.select_related('created_by').get(pk=int(edit_id_raw))
        except (OperationsExpense.DoesNotExist, ValueError, TypeError):
            edit_expense = None
            messages.error(request, 'Selected operation entry was not found.')

    if request.method == 'POST':
        expense_id = request.POST.get('expense_id')
        if expense_id:
            edit_expense = get_object_or_404(OperationsExpense, pk=expense_id)
            if not _can_manage_operation_expense(request.user, edit_expense):
                messages.error(request, 'You do not have permission to edit this operation entry.')
                return redirect(f"{request.path}?date={edit_expense.operation_date.isoformat()}")
            form = OperationsExpenseForm(request.POST, instance=edit_expense)
        else:
            form = OperationsExpenseForm(request.POST)

        if form.is_valid():
            expense = form.save(commit=False)
            if not expense.created_by_id:
                expense.created_by = request.user
            expense.save()
            # Persist any new detail value for future autocomplete suggestions
            detail_value = (expense.details or '').strip()
            if detail_value:
                _predefined_expense_details = [
                    'Salesperson 1 Salary', 'Salesperson 2 Salary', 'Salesperson 3 Salary',
                    'Other Salesperson Salary', 'Shortfall', 'Waste', 'Gift', 'Others',
                ]
                if detail_value not in _predefined_expense_details:
                    ExpenseDetailOption.objects.get_or_create(name=detail_value)
            if expense_id:
                messages.success(request, 'Operation expense entry updated successfully.')
            else:
                messages.success(request, 'Operation expense entry saved successfully.')
            return redirect(f"{request.path}?date={expense.operation_date.isoformat()}")
    else:
        if edit_expense:
            if not _can_manage_operation_expense(request.user, edit_expense):
                messages.error(request, 'You do not have permission to edit this operation entry.')
                return redirect(f"{request.path}?date={selected_date.isoformat()}")
            selected_date = edit_expense.operation_date
            form = OperationsExpenseForm(instance=edit_expense)
        else:
            form = OperationsExpenseForm(initial={'operation_date': selected_date})

    day_entries = OperationsExpense.objects.filter(operation_date=selected_date).select_related('created_by')
    total_day_operation_cost = day_entries.aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']

    day_revenue = Sales.objects.filter(sale_date=selected_date).aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']
    day_operation_income = OperationsIncome.objects.filter(income_date=selected_date).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']

    previous_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)

    _predefined_expense_details = [
        'Salesperson 1 Salary', 'Salesperson 2 Salary', 'Salesperson 3 Salary',
        'Other Salesperson Salary', 'Shortfall', 'Waste', 'Gift', 'Others',
    ]
    custom_options = list(ExpenseDetailOption.objects.values_list('name', flat=True))
    expense_detail_options = _predefined_expense_details + [
        o for o in custom_options if o not in _predefined_expense_details
    ]

    context = {
        'form': form,
        'selected_date': selected_date,
        'previous_date': previous_date,
        'next_date': next_date,
        'operations_entries': day_entries,
        'total_day_operation_cost': total_day_operation_cost,
        'day_revenue': day_revenue,
        'day_operation_income': day_operation_income,
        'day_net_after_operations': day_revenue + day_operation_income - total_day_operation_cost,
        'edit_expense': edit_expense,
        'expense_detail_options': expense_detail_options,
    }
    return render(request, 'inventory/quick_operations_entry.html', context)


@login_required
def quick_income_entry(request):
    """Quick entry for dated operational income."""
    selected_date_raw = request.GET.get('date')
    if selected_date_raw:
        try:
            selected_date = datetime.fromisoformat(selected_date_raw).date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()

    edit_income = None
    edit_id_raw = request.GET.get('edit')
    if edit_id_raw:
        try:
            edit_income = OperationsIncome.objects.select_related('created_by').get(pk=int(edit_id_raw))
        except (OperationsIncome.DoesNotExist, ValueError, TypeError):
            edit_income = None
            messages.error(request, 'Selected income entry was not found.')

    if request.method == 'POST':
        income_id = request.POST.get('income_id')
        if income_id:
            edit_income = get_object_or_404(OperationsIncome, pk=income_id)
            if not _can_manage_operation_income(request.user, edit_income):
                messages.error(request, 'You do not have permission to edit this income entry.')
                return redirect(f"{request.path}?date={edit_income.income_date.isoformat()}")
            form = OperationsIncomeForm(request.POST, instance=edit_income)
        else:
            form = OperationsIncomeForm(request.POST)

        if form.is_valid():
            income = form.save(commit=False)
            if not income.created_by_id:
                income.created_by = request.user
            income.save()
            if income_id:
                messages.success(request, 'Income entry updated successfully.')
            else:
                messages.success(request, 'Income entry saved successfully.')
            return redirect(f"{request.path}?date={income.income_date.isoformat()}")
    else:
        if edit_income:
            if not _can_manage_operation_income(request.user, edit_income):
                messages.error(request, 'You do not have permission to edit this income entry.')
                return redirect(f"{request.path}?date={selected_date.isoformat()}")
            selected_date = edit_income.income_date
            form = OperationsIncomeForm(instance=edit_income)
        else:
            form = OperationsIncomeForm(initial={'income_date': selected_date})

    income_entries = OperationsIncome.objects.filter(income_date=selected_date).select_related('created_by')
    total_day_income = income_entries.aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    total_day_operation_cost = OperationsExpense.objects.filter(operation_date=selected_date).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    day_revenue = Sales.objects.filter(sale_date=selected_date).aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']

    previous_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)

    context = {
        'form': form,
        'selected_date': selected_date,
        'previous_date': previous_date,
        'next_date': next_date,
        'income_entries': income_entries,
        'total_day_income': total_day_income,
        'total_day_operation_cost': total_day_operation_cost,
        'day_revenue': day_revenue,
        'day_net_after_operations': day_revenue + total_day_income - total_day_operation_cost,
        'edit_income': edit_income,
    }
    return render(request, 'inventory/quick_income_entry.html', context)


@login_required
def delete_operations_expense(request, expense_id):
    """Delete an operations expense entry from the quick operations sheet."""
    expense = get_object_or_404(OperationsExpense, pk=expense_id)
    selected_date = request.POST.get('selected_date') or request.GET.get('date') or expense.operation_date.isoformat()

    if request.method != 'POST':
        messages.error(request, 'Invalid request method for deleting operation entry.')
        return redirect(f"{reverse('quick_operations_entry')}?date={selected_date}")

    if not _can_manage_operation_expense(request.user, expense):
        messages.error(request, 'You do not have permission to delete this operation entry.')
        return redirect(f"{reverse('quick_operations_entry')}?date={selected_date}")

    expense.delete()
    messages.success(request, 'Operation expense entry deleted successfully.')
    return redirect(f"{reverse('quick_operations_entry')}?date={selected_date}")


@login_required
def delete_operation_income(request, income_id):
    """Delete an operations income entry from quick income sheet."""
    income = get_object_or_404(OperationsIncome, pk=income_id)
    selected_date = request.POST.get('selected_date') or request.GET.get('date') or income.income_date.isoformat()

    if request.method != 'POST':
        messages.error(request, 'Invalid request method for deleting income entry.')
        return redirect(f"{reverse('quick_income_entry')}?date={selected_date}")

    if not _can_manage_operation_income(request.user, income):
        messages.error(request, 'You do not have permission to delete this income entry.')
        return redirect(f"{reverse('quick_income_entry')}?date={selected_date}")

    income.delete()
    messages.success(request, 'Income entry deleted successfully.')
    return redirect(f"{reverse('quick_income_entry')}?date={selected_date}")


@login_required
def expenses_history(request):
    """Operations expenses history with optional date-range filter."""
    context = _build_expenses_history_context(request)
    return render(request, 'inventory/expenses_history.html', context)


def _build_expenses_history_context(request):
    """Build expenses history context using optional date-range filters."""

    start_date_raw = (request.GET.get('start_date') or '').strip()
    end_date_raw = (request.GET.get('end_date') or '').strip()
    details_filter = (request.GET.get('details') or '').strip()

    if not start_date_raw and not end_date_raw and not details_filter:
        _predefined = [
            'Salesperson 1 Salary', 'Salesperson 2 Salary', 'Salesperson 3 Salary',
            'Other Salesperson Salary', 'Shortfall', 'Waste', 'Gift', 'Others',
        ]
        _custom = list(ExpenseDetailOption.objects.values_list('name', flat=True))
        return {
            'entries': OperationsExpense.objects.none(),
            'start_date': '',
            'end_date': '',
            'details_filter': '',
            'total_operation_cost': 0,
            'no_filter': True,
            'expense_detail_options': _predefined + [o for o in _custom if o not in _predefined],
        }

    start_date = None
    end_date = None

    if start_date_raw:
        try:
            start_date = datetime.fromisoformat(start_date_raw).date()
        except ValueError:
            start_date = None

    if end_date_raw:
        try:
            end_date = datetime.fromisoformat(end_date_raw).date()
        except ValueError:
            end_date = None

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    entries = OperationsExpense.objects.all()
    if start_date:
        entries = entries.filter(operation_date__gte=start_date)
    if end_date:
        entries = entries.filter(operation_date__lte=end_date)
    if details_filter:
        entries = entries.filter(details__icontains=details_filter)

    entries = entries.select_related('created_by')

    total_operation_cost = entries.aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']

    _predefined_expense_details = [
        'Salesperson 1 Salary', 'Salesperson 2 Salary', 'Salesperson 3 Salary',
        'Other Salesperson Salary', 'Shortfall', 'Waste', 'Gift', 'Others',
    ]
    custom_options = list(ExpenseDetailOption.objects.values_list('name', flat=True))

    context = {
        'entries': entries,
        'start_date': start_date,
        'end_date': end_date,
        'details_filter': details_filter,
        'total_operation_cost': total_operation_cost,
        'no_filter': False,
        'expense_detail_options': _predefined_expense_details + [o for o in custom_options if o not in _predefined_expense_details],
    }
    return context


@login_required
def print_expenses_html(request):
    """Print-friendly HTML for expenses history."""
    context = _build_expenses_history_context(request)
    context['now'] = timezone.now()
    return render(request, 'inventory/print_expenses.html', context)


@login_required
def print_expenses_pdf(request):
    """Export expenses history as PDF."""
    context = _build_expenses_history_context(request)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        start_date = context['start_date']
        end_date = context['end_date']

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="expenses_{start_date}_{end_date}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.9 * inch, height=0.9 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.12 * inch))

        title = Paragraph(
            f"<b>Expenses Report - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph(
            f"<b>Total Operation Cost:</b> Rs.{context['total_operation_cost']:.2f}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.15 * inch))

        data = [['Date', 'Details of Operation', 'Amount', 'Recorded By', 'Created At']]
        for item in context['entries']:
            recorded_by = (
                item.created_by.get_full_name() or item.created_by.username
                if item.created_by else '-'
            )
            data.append([
                item.operation_date.strftime('%Y-%m-%d'),
                item.details,
                f"Rs.{item.amount:.2f}",
                recorded_by,
                timezone.localtime(item.created_at).strftime('%Y-%m-%d %H:%M'),
            ])

        table = Table(data, colWidths=[1.0 * inch, 2.2 * inch, 0.9 * inch, 1.3 * inch, 1.2 * inch], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.7, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        return response
    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('expenses_history')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('expenses_history')


@login_required
def print_expenses_excel(request):
    """Export expenses history as Excel."""
    context = _build_expenses_history_context(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        start_date = context['start_date']
        end_date = context['end_date']

        wb = Workbook()
        ws = wb.active
        ws.title = 'Expenses Report'

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 20

        ws['A1'] = f"EXPENSES REPORT - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = f"Total Operation Cost: Rs.{context['total_operation_cost']:.2f}"

        headers = ['Date', 'Details of Operation', 'Amount', 'Recorded By', 'Created At']
        header_row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row = header_row + 1
        for item in context['entries']:
            recorded_by = (
                item.created_by.get_full_name() or item.created_by.username
                if item.created_by else '-'
            )
            ws.cell(row=row, column=1).value = item.operation_date.strftime('%Y-%m-%d')
            ws.cell(row=row, column=2).value = item.details
            ws.cell(row=row, column=3).value = float(item.amount)
            ws.cell(row=row, column=4).value = recorded_by
            ws.cell(row=row, column=5).value = timezone.localtime(item.created_at).strftime('%Y-%m-%d %H:%M')
            ws.cell(row=row, column=3).number_format = '₹#,##0.00'
            row += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="expenses_{start_date}_{end_date}.xlsx"'
        wb.save(response)
        return response
    except ImportError as e:
        messages.error(request, f'Excel generation not available. Please install openpyxl: {str(e)}')
        return redirect('expenses_history')
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('expenses_history')


@login_required
def print_expenses_csv(request):
    """Export expenses history as CSV."""
    context = _build_expenses_history_context(request)

    start_date = context['start_date']
    end_date = context['end_date']

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="expenses_{start_date}_{end_date}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Expenses Report'])
    writer.writerow(['Start Date', start_date.strftime('%Y-%m-%d')])
    writer.writerow(['End Date', end_date.strftime('%Y-%m-%d')])
    writer.writerow(['Total Operation Cost', f"{context['total_operation_cost']:.2f}"])
    writer.writerow([])
    writer.writerow(['Date', 'Details of Operation', 'Amount', 'Recorded By', 'Created At'])

    for item in context['entries']:
        recorded_by = (
            item.created_by.get_full_name() or item.created_by.username
            if item.created_by else '-'
        )
        writer.writerow([
            item.operation_date.strftime('%Y-%m-%d'),
            item.details,
            f"{item.amount:.2f}",
            recorded_by,
            timezone.localtime(item.created_at).strftime('%Y-%m-%d %H:%M'),
        ])

    return response

# ==================== REPORTS MODULE ====================

@login_required
def reports_dashboard(request):
    """Main reports dashboard"""
    today = timezone.now().date()
    
    # Daily stats
    daily_revenue = Sales.objects.filter(sale_date=today).aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']
    
    daily_sales = Sales.objects.filter(sale_date=today).count()
    
    # Weekly stats
    last_week = today - timedelta(days=7)
    weekly_revenue = Sales.objects.filter(
        sale_date__gte=last_week,
        sale_date__lte=today
    ).aggregate(total=Coalesce(Sum('total_price'), 0, output_field=DecimalField()))['total']
    
    weekly_sales = Sales.objects.filter(
        sale_date__gte=last_week,
        sale_date__lte=today
    ).count()
    
    context = {
        'daily_revenue': daily_revenue,
        'daily_sales': daily_sales,
        'weekly_revenue': weekly_revenue,
        'weekly_sales': weekly_sales,
    }
    
    return render(request, 'inventory/reports_dashboard.html', context)


def _build_daily_report_context(selected_date):

    # Add total combined stock taken for the date
    stock_taken_qs = SalesStockTaken.objects.filter(sales_date=selected_date)
    total_combined_stock_taken = stock_taken_qs.aggregate(total=Coalesce(Sum('stock_taken_count'), 0))['total']

    sales_qs = Sales.objects.filter(sale_date=selected_date).select_related('product', 'recorded_by')
    grouped_sales = {}
    for sale in sales_qs:
        product_name = normalize_sales_product_name(sale.product.name)
        if product_name not in grouped_sales:
            grouped_sales[product_name] = {
                'product_name': product_name,
                'quantity': 0,
                'revenue': Decimal('0.0'),
                'cost': Decimal('0.0'),
                'profit': Decimal('0.0'),
                'recorded_by': set(),
            }

        grouped_sales[product_name]['quantity'] += sale.quantity
        grouped_sales[product_name]['revenue'] += sale.total_price
        grouped_sales[product_name]['cost'] += Decimal(sale.quantity) * sale.product.cost_price
        grouped_sales[product_name]['recorded_by'].add(
            sale.recorded_by.get_full_name() or sale.recorded_by.username
        )

    sales = []
    for item in grouped_sales.values():
        item['unit_price'] = item['revenue'] / item['quantity'] if item['quantity'] else Decimal('0.0')
        item['profit'] = item['revenue'] - item['cost']
        item['recorded_by'] = ', '.join(sorted(item['recorded_by'])) if item['recorded_by'] else '-'
        sales.append(item)

    sales.sort(
        key=lambda item: get_report_product_sort_key(item['product_name'])
    )

    total_revenue = sales_qs.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']
    total_cost = sum((item['cost'] for item in sales), Decimal('0.0'))
    total_profit = total_revenue - total_cost
    total_operation_cost = OperationsExpense.objects.filter(operation_date=selected_date).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    net_profit = total_profit - total_operation_cost

    # Calculate total stock as of selected date
    products = Product.objects.filter(is_active=True)
    stock_map = get_stock_as_of_date_map(products, selected_date)
    total_stock = sum(stock_map.values())

    return {
        'selected_date': selected_date,
        'sales': sales,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'total_operation_cost': total_operation_cost,
        'net_profit': net_profit,
        'total_transactions': len(sales),
        'total_stock': total_stock,
        'total_combined_stock_taken': total_combined_stock_taken,
    }


def _build_weekly_report_context(start_date, end_date, salesperson=None):
    sales = Sales.objects.filter(
        sale_date__gte=start_date,
        sale_date__lte=end_date
    ).select_related('product', 'recorded_by')
    if salesperson is not None:
        sales = sales.filter(recorded_by=salesperson)

    total_revenue = sales.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']

    total_cost = sum(Decimal(sale.quantity) * sale.product.cost_price for sale in sales)
    total_profit = total_revenue - total_cost

    daily_data = {}
    for i in range((end_date - start_date).days + 1):
        current_date = start_date + timedelta(days=i)
        daily_sales = sales.filter(sale_date=current_date)
        daily_data[current_date.strftime('%a, %m/%d')] = {
            'count': daily_sales.count(),
            'quantity': daily_sales.aggregate(
                total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
            )['total'],
            'revenue': daily_sales.aggregate(
                total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
            )['total']
        }

    weekly_product_breakdown_map = {}
    for sale in sales:
        product_name = normalize_sales_product_name(sale.product.name)
        if product_name not in weekly_product_breakdown_map:
            weekly_product_breakdown_map[product_name] = {
                'product_name': product_name,
                'sort_sku': sale.product.sku or '',
                'quantity': 0,
                'revenue': Decimal('0.0'),
                'cost': Decimal('0.0'),
                'profit': Decimal('0.0'),
            }
        else:
            current_sku = weekly_product_breakdown_map[product_name]['sort_sku']
            candidate_sku = sale.product.sku or ''
            if candidate_sku and (not current_sku or candidate_sku < current_sku):
                weekly_product_breakdown_map[product_name]['sort_sku'] = candidate_sku

        weekly_product_breakdown_map[product_name]['quantity'] += sale.quantity
        weekly_product_breakdown_map[product_name]['revenue'] += sale.total_price
        weekly_product_breakdown_map[product_name]['cost'] += Decimal(sale.quantity) * sale.product.cost_price

    weekly_product_breakdown = []
    for item in weekly_product_breakdown_map.values():
        item['unit_price'] = item['revenue'] / item['quantity'] if item['quantity'] else Decimal('0.0')
        item['profit'] = item['revenue'] - item['cost']
        item['margin'] = (item['profit'] / item['revenue']) * 100 if item['revenue'] else Decimal('0.0')
        weekly_product_breakdown.append(item)

    weekly_product_breakdown.sort(
        key=lambda item: get_report_product_sort_key(item['product_name'], item.get('sort_sku', ''))
    )

    return {
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'total_transactions': sales.count(),
        'daily_data': daily_data,
        'weekly_product_breakdown': weekly_product_breakdown,
    }


def _get_weekly_salesperson_filter_data(request):
    """Return salespeople options and selected salesperson from request query params."""
    selected_salesperson = None
    selected_salesperson_id = (request.GET.get('salesperson') or '').strip()
    salespeople = User.objects.filter(
        is_active=True,
        sales__isnull=False,
    ).distinct().order_by('first_name', 'username')

    if selected_salesperson_id:
        try:
            selected_salesperson = salespeople.get(pk=int(selected_salesperson_id))
        except (User.DoesNotExist, ValueError, TypeError):
            selected_salesperson = None
            selected_salesperson_id = ''
            messages.warning(request, 'Selected salesperson not found. Showing all salespersons.')

    return salespeople, selected_salesperson, selected_salesperson_id


def _get_current_week_date_range():
    """Return current week date range (Monday to Sunday) based on today's date."""
    today = timezone.now().date()
    start_date = today - timedelta(days=today.weekday())
    end_date = start_date + timedelta(days=6)
    return start_date, end_date


def _build_profit_report_context(start_date, end_date):
    sales = Sales.objects.filter(
        sale_date__gte=start_date,
        sale_date__lte=end_date
    ).select_related('product')

    products_profit = {}
    for sale in sales:
        product_name = normalize_sales_product_name(sale.product.name)
        profit = sale.get_profit()
        if product_name not in products_profit:
            products_profit[product_name] = {'profit': Decimal('0.0'), 'revenue': Decimal('0.0'), 'cost': Decimal('0.0'), 'quantity': 0}
        products_profit[product_name]['profit'] += Decimal(profit)
        products_profit[product_name]['revenue'] += sale.total_price
        products_profit[product_name]['cost'] += Decimal(sale.quantity) * sale.product.cost_price
        products_profit[product_name]['quantity'] += sale.quantity

    for product in products_profit.values():
        if product['revenue'] > 0:
            product['margin'] = (product['profit'] / product['revenue']) * 100
        else:
            product['margin'] = Decimal('0.0')

    sorted_products = sorted(products_profit.items(), key=lambda x: x[1]['profit'], reverse=True)

    total_revenue = sales.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']

    total_cost = sum(Decimal(sale.quantity) * sale.product.cost_price for sale in sales)
    total_profit = total_revenue - total_cost

    total_quantity = sales.aggregate(
        total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
    )['total']

    return {
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'total_quantity': total_quantity,
        'products_profit': sorted_products,
    }


def _build_income_statement_context(start_date, end_date):
    sales = Sales.objects.filter(
        sale_date__gte=start_date,
        sale_date__lte=end_date,
    ).select_related('product')

    operating_income_entries = OperationsIncome.objects.filter(
        income_date__gte=start_date,
        income_date__lte=end_date,
    ).select_related('created_by').order_by('-income_date', '-created_at')

    operating_expense_entries = OperationsExpense.objects.filter(
        operation_date__gte=start_date,
        operation_date__lte=end_date,
    ).select_related('created_by').order_by('-operation_date', '-created_at')

    sales_revenue = sales.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']

    cost_of_goods_sold = sum(
        Decimal(sale.quantity) * sale.product.cost_price for sale in sales
    )

    gross_profit = sales_revenue - cost_of_goods_sold

    operating_income_total = operating_income_entries.aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']

    operating_expense_total = operating_expense_entries.aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']

    net_profit_before_tax = gross_profit + operating_income_total - operating_expense_total

    sales_by_product = {}
    for sale in sales:
        product_name = normalize_sales_product_name(sale.product.name)
        if product_name not in sales_by_product:
            sales_by_product[product_name] = {
                'product_name': product_name,
                'quantity': 0,
                'revenue': Decimal('0.0'),
                'cost': Decimal('0.0'),
            }

        sales_by_product[product_name]['quantity'] += sale.quantity
        sales_by_product[product_name]['revenue'] += sale.total_price
        sales_by_product[product_name]['cost'] += Decimal(sale.quantity) * sale.product.cost_price

    sales_breakdown = []
    for row in sales_by_product.values():
        row['gross_profit'] = row['revenue'] - row['cost']
        sales_breakdown.append(row)

    sales_breakdown.sort(key=lambda row: row['product_name'])

    gross_margin = (gross_profit / sales_revenue) * 100 if sales_revenue else Decimal('0.0')
    net_margin = (net_profit_before_tax / sales_revenue) * 100 if sales_revenue else Decimal('0.0')

    return {
        'start_date': start_date,
        'end_date': end_date,
        'sales_revenue': sales_revenue,
        'cost_of_goods_sold': cost_of_goods_sold,
        'gross_profit': gross_profit,
        'operating_income_total': operating_income_total,
        'operating_expense_total': operating_expense_total,
        'net_profit_before_tax': net_profit_before_tax,
        'gross_margin': gross_margin,
        'net_margin': net_margin,
        'sales_breakdown': sales_breakdown,
        'operating_income_entries': operating_income_entries,
        'operating_expense_entries': operating_expense_entries,
        'sales_transaction_count': sales.count(),
    }


def _extract_positive_adjustment_qty(notes):
    if not notes:
        return 0
    match = re.search(r'Adjustment mode:\s*\+(\d+)', notes, flags=re.IGNORECASE)
    return int(match.group(1)) if match else 0


def _build_stock_report_context(start_date, end_date, include_positive_adjustments=False, report_mode='detailed'):
    movement_types = ['IN']
    if include_positive_adjustments:
        movement_types.append('ADJUSTMENT')

    raw_movements = Inventory.objects.filter(
        movement_type__in=movement_types,
        movement_date__gte=start_date,
        movement_date__lte=end_date,
        product__category__in=['Indian Kulfi', 'Kulfi Corner'],
    ).select_related('product', 'created_by').order_by(
        '-movement_date',
        'product__category',
        'product__sku',
        '-created_at',
    )

    movement_rows = []
    total_quantity = 0
    indian_kulfi_quantity = 0
    kulfi_corner_quantity = 0
    total_purchase_cost = Decimal('0.0')
    grouped_general_rows = {}

    for movement in raw_movements:
        if movement.movement_type == 'IN':
            qty_in = movement.quantity
            entry_type = 'Stock In'
        else:
            qty_in = _extract_positive_adjustment_qty(movement.notes)
            if qty_in <= 0:
                continue
            entry_type = 'Positive Adjustment'

        # Use manufacturer stamped at entry time if present; fall back to cost-price detection.
        resolved_manufacturer = (
            _extract_manufacturer_from_notes(movement.notes)
            or _identify_manufacturer_from_cost(movement.product.cost_price)
        )

        unit_cost_val = movement.unit_cost or movement.product.cost_price
        movement_rows.append({
            'movement': movement,
            'qty_in': qty_in,
            'entry_type': entry_type,
            'resolved_manufacturer': resolved_manufacturer,
            'unit_cost': unit_cost_val,
            'total_cost': Decimal(qty_in) * unit_cost_val,
        })

        total_quantity += qty_in
        total_purchase_cost += Decimal(qty_in) * unit_cost_val
        if resolved_manufacturer == 'Indian Kulfi':
            indian_kulfi_quantity += qty_in
        elif resolved_manufacturer == 'Kulfi Corner':
            kulfi_corner_quantity += qty_in

        general_key = (
            movement.movement_date,
            resolved_manufacturer,
        )
        if general_key not in grouped_general_rows:
            grouped_general_rows[general_key] = {
                'movement_date': movement.movement_date,
                'manufacturer': resolved_manufacturer,
                'total_quantity': 0,
                'overall_cost_price': Decimal('0.0'),
            }

        grouped_general_rows[general_key]['total_quantity'] += qty_in
        grouped_general_rows[general_key]['overall_cost_price'] += Decimal(qty_in) * (movement.unit_cost or movement.product.cost_price)

    general_rows = []
    for grouped_row in grouped_general_rows.values():
        row_total_qty = grouped_row['total_quantity']
        general_rows.append({
            'movement_date': grouped_row['movement_date'],
            'manufacturer': grouped_row['manufacturer'],
            'total_quantity': row_total_qty,
            'total_packs': (Decimal(row_total_qty) / Decimal('6')) if row_total_qty else Decimal('0.0'),
            'overall_cost_price': grouped_row['overall_cost_price'],
        })

    general_rows.sort(key=lambda row: (
        -row['movement_date'].toordinal(),
        row['manufacturer'],
    ))

    return {
        'start_date': start_date,
        'end_date': end_date,
        'movement_rows': movement_rows,
        'total_entries': len(movement_rows),
        'total_quantity': total_quantity,
        'indian_kulfi_quantity': indian_kulfi_quantity,
        'kulfi_corner_quantity': kulfi_corner_quantity,
        'total_purchase_cost': total_purchase_cost,
        'include_positive_adjustments': include_positive_adjustments,
        'general_rows': general_rows,
        'report_mode': report_mode,
    }

@login_required
def daily_report(request):
    """Daily sales report"""
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    try:
        selected_date = datetime.fromisoformat(selected_date).date()
    except ValueError:
        selected_date = timezone.now().date()

    context = _build_daily_report_context(selected_date)
    
    return render(request, 'inventory/daily_report.html', context)

@login_required
def weekly_report(request):
    """Weekly sales report"""
    start_date_raw = (request.GET.get('start_date') or '').strip()
    end_date_raw = (request.GET.get('end_date') or '').strip()

    salespeople, selected_salesperson, selected_salesperson_id = _get_weekly_salesperson_filter_data(request)

    if not start_date_raw and not end_date_raw:
        context = {
            'start_date': '',
            'end_date': '',
            'total_transactions': 0,
            'total_revenue': 0,
            'total_cost': 0,
            'total_profit': 0,
            'daily_data': OrderedDict(),
            'weekly_product_breakdown': [],
            'no_filter': True,
        }
    else:
        start_date = None
        end_date = None

        if start_date_raw:
            try:
                start_date = datetime.fromisoformat(start_date_raw).date()
            except ValueError:
                start_date = None

        if end_date_raw:
            try:
                end_date = datetime.fromisoformat(end_date_raw).date()
            except ValueError:
                end_date = None

        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date

        if start_date is None and end_date is not None:
            start_date = end_date - timedelta(days=6)
        if end_date is None and start_date is not None:
            end_date = start_date + timedelta(days=6)

        context = _build_weekly_report_context(start_date, end_date, salesperson=selected_salesperson)
        context['no_filter'] = False

    context.update({
        'salespeople': salespeople,
        'selected_salesperson': selected_salesperson,
        'selected_salesperson_id': selected_salesperson_id,
    })
    
    return render(request, 'inventory/weekly_report.html', context)

@login_required
def profit_report(request):
    """Profit analysis report"""
    start_date_raw = (request.GET.get('start_date') or '').strip()
    end_date_raw = (request.GET.get('end_date') or '').strip()

    if not start_date_raw and not end_date_raw:
        context = {
            'start_date': '',
            'end_date': '',
            'products_profit': [],
            'total_quantity': 0,
            'total_revenue': 0,
            'total_cost': 0,
            'total_profit': 0,
            'no_filter': True,
        }
    else:
        start_date = None
        end_date = None

        if start_date_raw:
            try:
                start_date = datetime.fromisoformat(start_date_raw).date()
            except ValueError:
                start_date = None

        if end_date_raw:
            try:
                end_date = datetime.fromisoformat(end_date_raw).date()
            except ValueError:
                end_date = None

        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date

        if start_date is None and end_date is not None:
            start_date = end_date - timedelta(days=6)
        if end_date is None and start_date is not None:
            end_date = start_date + timedelta(days=6)

        context = _build_profit_report_context(start_date, end_date)
        context['no_filter'] = False

    return render(request, 'inventory/profit_report.html', context)


@login_required
def income_statement(request):
    """Income Statement (Profit and Loss Statement) for a selected date range."""
    start_date_raw = (request.GET.get('start_date') or '').strip()
    end_date_raw = (request.GET.get('end_date') or '').strip()

    if not start_date_raw and not end_date_raw:
        context = {
            'start_date': '',
            'end_date': '',
            'sales_revenue': 0,
            'gross_profit': 0,
            'gross_margin': 0,
            'operating_expense_total': 0,
            'operating_income_total': 0,
            'net_profit_before_tax': 0,
            'net_margin': 0,
            'cost_of_goods_sold': 0,
            'sales_breakdown': [],
            'operating_income_entries': [],
            'operating_expense_entries': [],
            'no_filter': True,
        }
    else:
        start_date = None
        end_date = None

        if start_date_raw:
            try:
                start_date = datetime.fromisoformat(start_date_raw).date()
            except ValueError:
                start_date = None
        if end_date_raw:
            try:
                end_date = datetime.fromisoformat(end_date_raw).date()
            except ValueError:
                end_date = None

        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date
            messages.info(request, 'Start date and end date were swapped to apply a valid date range.')

        if start_date is None and end_date is not None:
            start_date = end_date - timedelta(days=6)
        if end_date is None and start_date is not None:
            end_date = start_date + timedelta(days=6)

        context = _build_income_statement_context(start_date, end_date)
        context['no_filter'] = False

    return render(request, 'inventory/income_statement.html', context)


@login_required
def stock_report(request):
    """Stock-in report by date range for Indian Kulfi and Kulfi Corner."""
    start_date_raw = (request.GET.get('start_date') or '').strip()
    end_date_raw = (request.GET.get('end_date') or '').strip()
    include_positive_adjustments = request.GET.get('include_adjustments') == '1'
    report_mode = request.GET.get('view_mode', 'detailed')
    if report_mode not in ('general', 'detailed'):
        report_mode = 'detailed'

    if not start_date_raw and not end_date_raw:
        context = {
            'start_date': '',
            'end_date': '',
            'movement_rows': [],
            'general_rows': [],
            'total_entries': 0,
            'total_quantity': 0,
            'indian_kulfi_quantity': 0,
            'kulfi_corner_quantity': 0,
            'total_purchase_cost': Decimal('0.0'),
            'include_positive_adjustments': include_positive_adjustments,
            'report_mode': report_mode,
            'no_filter': True,
        }
        return render(request, 'inventory/stock_report.html', context)

    start_date = None
    end_date = None

    if start_date_raw:
        try:
            start_date = datetime.fromisoformat(start_date_raw).date()
        except ValueError:
            start_date = None
    if end_date_raw:
        try:
            end_date = datetime.fromisoformat(end_date_raw).date()
        except ValueError:
            end_date = None

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date
        messages.info(request, 'Start date and end date were swapped to apply a valid date range.')

    if start_date is None and end_date is not None:
        start_date = end_date - timedelta(days=6)
    if end_date is None and start_date is not None:
        end_date = start_date + timedelta(days=6)

    context = _build_stock_report_context(
        start_date,
        end_date,
        include_positive_adjustments=include_positive_adjustments,
        report_mode=report_mode,
    )
    context['no_filter'] = False
    return render(request, 'inventory/stock_report.html', context)


@login_required
def print_stock_report_html(request):
    today = timezone.now().date()
    start_date_raw = request.GET.get('start_date')
    end_date_raw = request.GET.get('end_date')
    include_positive_adjustments = request.GET.get('include_adjustments') == '1'
    report_mode = request.GET.get('view_mode', 'detailed')
    if report_mode not in ('general', 'detailed'):
        report_mode = 'detailed'

    if not start_date_raw:
        start_date = today - timedelta(days=30)
    else:
        try:
            start_date = datetime.fromisoformat(start_date_raw).date()
        except ValueError:
            start_date = today - timedelta(days=30)

    if not end_date_raw:
        end_date = today
    else:
        try:
            end_date = datetime.fromisoformat(end_date_raw).date()
        except ValueError:
            end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    context = _build_stock_report_context(
        start_date,
        end_date,
        include_positive_adjustments=include_positive_adjustments,
        report_mode=report_mode,
    )
    context['now'] = timezone.now()
    return render(request, 'inventory/print_stock_report.html', context)


@login_required
def print_stock_report_pdf(request):
    today = timezone.now().date()
    start_date_raw = request.GET.get('start_date')
    end_date_raw = request.GET.get('end_date')
    include_positive_adjustments = request.GET.get('include_adjustments') == '1'
    report_mode = request.GET.get('view_mode', 'detailed')
    if report_mode not in ('general', 'detailed'):
        report_mode = 'detailed'

    if not start_date_raw:
        start_date = today - timedelta(days=30)
    else:
        try:
            start_date = datetime.fromisoformat(start_date_raw).date()
        except ValueError:
            start_date = today - timedelta(days=30)

    if not end_date_raw:
        end_date = today
    else:
        try:
            end_date = datetime.fromisoformat(end_date_raw).date()
        except ValueError:
            end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    context = _build_stock_report_context(
        start_date,
        end_date,
        include_positive_adjustments=include_positive_adjustments,
        report_mode=report_mode,
    )

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="stock_report_{start_date}_{end_date}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.9 * inch, height=0.9 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.12 * inch))

        title = Paragraph(
            f"<b>Stock Report - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        summary = (
            f"<b>Total Entries:</b> {context['total_entries']} | "
            f"<b>Total Qty In:</b> {context['total_quantity']} | "
            f"<b>Indian Kulfi:</b> {context['indian_kulfi_quantity']} | "
            f"<b>Kulfi Corner:</b> {context['kulfi_corner_quantity']} | "
            f"<b>Total Purchase Cost:</b> Rs.{context['total_purchase_cost']:.2f}"
        )
        elements.append(Paragraph(summary, styles['Normal']))
        if include_positive_adjustments:
            elements.append(Paragraph('<b>Included:</b> Positive adjustments', styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        if report_mode == 'general':
            data = [['Stock In Date', 'Total Packs', 'Total Quantity', 'Overall Cost Price', 'Manufacturer']]
            for row in context['general_rows']:
                data.append([
                    row['movement_date'].strftime('%Y-%m-%d'),
                    f"{row['total_packs']:.2f}",
                    str(row['total_quantity']),
                    f"Rs.{row['overall_cost_price']:.2f}",
                    row['manufacturer'],
                ])
            col_widths = [1.05 * inch, 1.0 * inch, 1.0 * inch, 1.4 * inch, 1.25 * inch]
        else:
            data = [['Date', 'Manufacturer', 'Entry Type', 'SKU', 'Product', 'Qty In']]
            for row in context['movement_rows']:
                movement = row['movement']
                data.append([
                    movement.movement_date.strftime('%Y-%m-%d'),
                    movement.product.category,
                    row['entry_type'],
                    movement.product.sku,
                    movement.product.name,
                    str(row['qty_in']),
                ])
            col_widths = [0.95 * inch, 1.0 * inch, 1.1 * inch, 0.9 * inch, 2.0 * inch, 0.7 * inch]

        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        if report_mode == 'general':
            table.setStyle(TableStyle([
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),
            ]))
        else:
            table.setStyle(TableStyle([
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),
            ]))
        elements.append(table)

        doc.build(elements)
        return response
    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('stock_report')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('stock_report')


@login_required
def print_stock_report_excel(request):
    today = timezone.now().date()
    start_date_raw = request.GET.get('start_date')
    end_date_raw = request.GET.get('end_date')
    include_positive_adjustments = request.GET.get('include_adjustments') == '1'
    report_mode = request.GET.get('view_mode', 'detailed')
    if report_mode not in ('general', 'detailed'):
        report_mode = 'detailed'

    if not start_date_raw:
        start_date = today - timedelta(days=30)
    else:
        try:
            start_date = datetime.fromisoformat(start_date_raw).date()
        except ValueError:
            start_date = today - timedelta(days=30)

    if not end_date_raw:
        end_date = today
    else:
        try:
            end_date = datetime.fromisoformat(end_date_raw).date()
        except ValueError:
            end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    context = _build_stock_report_context(
        start_date,
        end_date,
        include_positive_adjustments=include_positive_adjustments,
        report_mode=report_mode,
    )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Report"

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 26
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18

        ws['A1'] = f"STOCK REPORT - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = f"Total Entries: {context['total_entries']}"
        ws['A4'] = f"Total Qty In: {context['total_quantity']}"
        ws['A5'] = f"Indian Kulfi Qty: {context['indian_kulfi_quantity']}"
        ws['A6'] = f"Kulfi Corner Qty: {context['kulfi_corner_quantity']}"
        ws['A7'] = f"Total Purchase Cost: Rs.{context['total_purchase_cost']:.2f}"
        ws['A8'] = f"Positive Adjustments Included: {'Yes' if include_positive_adjustments else 'No'}"
        ws['A9'] = f"Report Mode: {'General' if report_mode == 'general' else 'Detailed'}"

        header_row = 11
        if report_mode == 'general':
            headers = ['Stock In Date', 'Total Packs', 'Total Quantity', 'Overall Cost Price', 'Manufacturer']
        else:
            headers = ['Date', 'Manufacturer', 'Entry Type', 'SKU', 'Product', 'Qty In']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row_number = header_row + 1
        if report_mode == 'general':
            for row in context['general_rows']:
                ws.cell(row=row_number, column=1).value = row['movement_date'].strftime('%Y-%m-%d')
                ws.cell(row=row_number, column=2).value = float(row['total_packs'])
                ws.cell(row=row_number, column=3).value = int(row['total_quantity'])
                ws.cell(row=row_number, column=4).value = float(row['overall_cost_price'])
                ws.cell(row=row_number, column=5).value = row['manufacturer']
                row_number += 1
        else:
            for row in context['movement_rows']:
                movement = row['movement']
                ws.cell(row=row_number, column=1).value = movement.movement_date.strftime('%Y-%m-%d')
                ws.cell(row=row_number, column=2).value = movement.product.category
                ws.cell(row=row_number, column=3).value = row['entry_type']
                ws.cell(row=row_number, column=4).value = movement.product.sku
                ws.cell(row=row_number, column=5).value = movement.product.name
                ws.cell(row=row_number, column=6).value = int(row['qty_in'])
                row_number += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="stock_report_{start_date}_{end_date}.xlsx"'
        wb.save(response)
        return response
    except ImportError as e:
        messages.error(request, f'Excel generation not available. Please install openpyxl: {str(e)}')
        return redirect('stock_report')
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('stock_report')


@login_required
def print_daily_report_html(request):
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    try:
        selected_date = datetime.fromisoformat(selected_date).date()
    except ValueError:
        selected_date = timezone.now().date()

    context = _build_daily_report_context(selected_date)
    context['now'] = timezone.now()
    return render(request, 'inventory/print_daily_report.html', context)


@login_required
def print_daily_report_pdf(request):
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    try:
        selected_date = datetime.fromisoformat(selected_date).date()
    except ValueError:
        selected_date = timezone.now().date()

    context = _build_daily_report_context(selected_date)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="daily_report_{selected_date}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.9 * inch, height=0.9 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.12 * inch))

        title = Paragraph(f"<b>Daily Report - {selected_date.strftime('%d %B %Y')}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.25 * inch))

        summary = (
            f"<b>Transactions:</b> {context['total_transactions']} | "
            f"<b>Revenue:</b> Rs.{context['total_revenue']:.2f} | "
            f"<b>Cost:</b> Rs.{context['total_cost']:.2f} | "
            f"<b>Profit:</b> Rs.{context['total_profit']:.2f} | "
            f"<b>Operation Cost:</b> Rs.{context['total_operation_cost']:.2f} | "
            f"<b>Net Profit:</b> Rs.{context['net_profit']:.2f}"
        )
        elements.append(Paragraph(summary, styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        data = [['Product', 'Qty', 'Unit Price', 'Revenue', 'Cost', 'Profit']]
        for item in context['sales']:
            data.append([
                item['product_name'],
                str(item['quantity']),
                f"Rs.{item['unit_price']:.2f}",
                f"Rs.{item['revenue']:.2f}",
                f"Rs.{item['cost']:.2f}",
                f"Rs.{item['profit']:.2f}",
            ])

        data.append([
            'TOTAL',
            '',
            '',
            f"Rs.{context['total_revenue']:.2f}",
            f"Rs.{context['total_cost']:.2f}",
            f"Rs.{context['total_profit']:.2f}",
        ])

        table = Table(data, colWidths=[2.2 * inch, 0.7 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        return response
    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('daily_report')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('daily_report')


@login_required
def print_daily_report_excel(request):
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    try:
        selected_date = datetime.fromisoformat(selected_date).date()
    except ValueError:
        selected_date = timezone.now().date()

    context = _build_daily_report_context(selected_date)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14

        ws['A1'] = f"DAILY REPORT - {selected_date.strftime('%d %B %Y')}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = f"Transactions: {context['total_transactions']}"
        ws['A4'] = f"Revenue: {context['total_revenue']:.2f}"
        ws['A5'] = f"Cost: {context['total_cost']:.2f}"
        ws['A6'] = f"Profit: {context['total_profit']:.2f}"
        ws['A7'] = f"Operation Cost: {context['total_operation_cost']:.2f}"
        ws['A8'] = f"Net Profit: {context['net_profit']:.2f}"

        headers = ['Product', 'Quantity', 'Unit Price', 'Revenue', 'Cost', 'Profit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row = 11
        for item in context['sales']:
            ws.cell(row=row, column=1).value = item['product_name']
            ws.cell(row=row, column=2).value = int(item['quantity'])
            ws.cell(row=row, column=3).value = float(item['unit_price'])
            ws.cell(row=row, column=4).value = float(item['revenue'])
            ws.cell(row=row, column=5).value = float(item['cost'])
            ws.cell(row=row, column=6).value = float(item['profit'])
            row += 1

        ws.cell(row=row, column=1).value = 'TOTAL'
        ws.cell(row=row, column=4).value = float(context['total_revenue'])
        ws.cell(row=row, column=5).value = float(context['total_cost'])
        ws.cell(row=row, column=6).value = float(context['total_profit'])
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = Font(bold=True)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="daily_report_{selected_date}.xlsx"'
        wb.save(response)
        return response
    except ImportError as e:
        messages.error(request, f'Excel generation not available. Please install openpyxl: {str(e)}')
        return redirect('daily_report')
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('daily_report')


@login_required
def print_weekly_report_html(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = today - timedelta(days=today.weekday())
    else:
        start_date = datetime.fromisoformat(start_date).date()

    if not end_date:
        end_date = start_date + timedelta(days=6)
    else:
        end_date = datetime.fromisoformat(end_date).date()

    _, selected_salesperson, selected_salesperson_id = _get_weekly_salesperson_filter_data(request)
    context = _build_weekly_report_context(start_date, end_date, salesperson=selected_salesperson)
    context['selected_salesperson'] = selected_salesperson
    context['selected_salesperson_id'] = selected_salesperson_id
    context['now'] = timezone.now()
    return render(request, 'inventory/print_weekly_report.html', context)


@login_required
def print_weekly_report_pdf(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = today - timedelta(days=today.weekday())
    else:
        start_date = datetime.fromisoformat(start_date).date()

    if not end_date:
        end_date = start_date + timedelta(days=6)
    else:
        end_date = datetime.fromisoformat(end_date).date()

    _, selected_salesperson, _ = _get_weekly_salesperson_filter_data(request)
    context = _build_weekly_report_context(start_date, end_date, salesperson=selected_salesperson)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="weekly_report_{start_date}_{end_date}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.9 * inch, height=0.9 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.12 * inch))

        title = Paragraph(
            f"<b>Weekly Report - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 0.25 * inch))

        summary = (
            f"<b>Transactions:</b> {context['total_transactions']} | "
            f"<b>Revenue:</b> Rs.{context['total_revenue']:.2f} | "
            f"<b>Cost:</b> Rs.{context['total_cost']:.2f} | "
            f"<b>Profit:</b> Rs.{context['total_profit']:.2f}"
        )
        if selected_salesperson is not None:
            salesperson_name = selected_salesperson.get_full_name() or selected_salesperson.username
            summary = f"{summary} | <b>Salesperson:</b> {salesperson_name}"
        elements.append(Paragraph(summary, styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph('<b>Daily Breakdown</b>', styles['Heading3']))
        daily_data = [['Date', 'Transactions', 'Quantity Sold', 'Revenue']]
        for day, row in context['daily_data'].items():
            daily_data.append([
                day,
                str(row['count']),
                f"{Decimal(row['quantity']):.0f}",
                f"Rs.{row['revenue']:.2f}",
            ])

        daily_table = Table(daily_data, colWidths=[1.6 * inch, 1.0 * inch, 1.2 * inch, 1.3 * inch])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(daily_table)
        elements.append(Spacer(1, 0.18 * inch))

        elements.append(Paragraph('<b>Product-wise Breakdown</b>', styles['Heading3']))

        data = [['Product', 'Qty', 'Revenue', 'Cost', 'Profit', 'Margin %']]
        for item in context['weekly_product_breakdown']:
            data.append([
                item['product_name'],
                str(item['quantity']),
                f"Rs.{item['revenue']:.2f}",
                f"Rs.{item['cost']:.2f}",
                f"Rs.{item['profit']:.2f}",
                f"{item['margin']:.2f}%",
            ])

        table = Table(data, colWidths=[2.2 * inch, 0.7 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 0.8 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        return response
    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('weekly_report')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('weekly_report')


@login_required
def print_weekly_report_excel(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = today - timedelta(days=today.weekday())
    else:
        start_date = datetime.fromisoformat(start_date).date()

    if not end_date:
        end_date = start_date + timedelta(days=6)
    else:
        end_date = datetime.fromisoformat(end_date).date()

    _, selected_salesperson, _ = _get_weekly_salesperson_filter_data(request)
    context = _build_weekly_report_context(start_date, end_date, salesperson=selected_salesperson)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Report"

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12

        ws['A1'] = f"WEEKLY REPORT - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
        ws['A1'].font = Font(bold=True, size=14)

        if selected_salesperson is not None:
            salesperson_name = selected_salesperson.get_full_name() or selected_salesperson.username
            ws['A2'] = f"Salesperson: {salesperson_name}"

        ws['A3'] = f"Transactions: {context['total_transactions']}"
        ws['A4'] = f"Revenue: {context['total_revenue']:.2f}"
        ws['A5'] = f"Cost: {context['total_cost']:.2f}"
        ws['A6'] = f"Profit: {context['total_profit']:.2f}"

        ws['A8'] = 'Daily Breakdown'
        ws['A8'].font = Font(bold=True)

        daily_headers = ['Date', 'Transactions', 'Quantity Sold', 'Revenue']
        daily_header_row = 9
        for col, header in enumerate(daily_headers, 1):
            cell = ws.cell(row=daily_header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row = daily_header_row + 1
        for day, data in context['daily_data'].items():
            ws.cell(row=row, column=1).value = day
            ws.cell(row=row, column=2).value = int(data['count'])
            ws.cell(row=row, column=3).value = int(data['quantity'])
            ws.cell(row=row, column=4).value = float(data['revenue'])
            row += 1

        row += 1
        ws.cell(row=row, column=1).value = 'Product-wise Breakdown'
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        headers = ['Product', 'Quantity', 'Revenue', 'Cost', 'Profit', 'Margin %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1
        for item in context['weekly_product_breakdown']:
            ws.cell(row=row, column=1).value = item['product_name']
            ws.cell(row=row, column=2).value = int(item['quantity'])
            ws.cell(row=row, column=3).value = float(item['revenue'])
            ws.cell(row=row, column=4).value = float(item['cost'])
            ws.cell(row=row, column=5).value = float(item['profit'])
            ws.cell(row=row, column=6).value = float(item['margin'])
            row += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="weekly_report_{start_date}_{end_date}.xlsx"'
        wb.save(response)
        return response
    except ImportError as e:
        messages.error(request, f'Excel generation not available. Please install openpyxl: {str(e)}')
        return redirect('weekly_report')
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('weekly_report')


@login_required
def print_profit_report_html(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = today - timedelta(days=30)
    else:
        start_date = datetime.fromisoformat(start_date).date()

    if not end_date:
        end_date = today
    else:
        end_date = datetime.fromisoformat(end_date).date()

    context = _build_profit_report_context(start_date, end_date)
    context['now'] = timezone.now()
    return render(request, 'inventory/print_profit_report.html', context)


@login_required
def print_profit_report_pdf(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = today - timedelta(days=30)
    else:
        start_date = datetime.fromisoformat(start_date).date()

    if not end_date:
        end_date = today
    else:
        end_date = datetime.fromisoformat(end_date).date()

    context = _build_profit_report_context(start_date, end_date)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="profit_report_{start_date}_{end_date}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.9 * inch, height=0.9 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.12 * inch))

        title = Paragraph(
            f"<b>Profit Report - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 0.25 * inch))

        summary = (
            f"<b>Revenue:</b> Rs.{context['total_revenue']:.2f} | "
            f"<b>Cost:</b> Rs.{context['total_cost']:.2f} | "
            f"<b>Profit:</b> Rs.{context['total_profit']:.2f}"
        )
        elements.append(Paragraph(summary, styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        data = [['Product', 'Qty', 'Revenue', 'Cost', 'Profit', 'Margin %']]
        for product_name, item in context['products_profit']:
            data.append([
                product_name,
                str(item['quantity']),
                f"Rs.{item['revenue']:.2f}",
                f"Rs.{item['cost']:.2f}",
                f"Rs.{item['profit']:.2f}",
                f"{item['margin']:.2f}%",
            ])

        table = Table(data, colWidths=[2.2 * inch, 0.7 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 0.8 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        return response
    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('profit_report')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('profit_report')


@login_required
def print_profit_report_excel(request):
    today = timezone.now().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = today - timedelta(days=30)
    else:
        start_date = datetime.fromisoformat(start_date).date()

    if not end_date:
        end_date = today
    else:
        end_date = datetime.fromisoformat(end_date).date()

    context = _build_profit_report_context(start_date, end_date)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Profit Report"

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12

        ws['A1'] = f"PROFIT REPORT - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = f"Revenue: {context['total_revenue']:.2f}"
        ws['A4'] = f"Cost: {context['total_cost']:.2f}"
        ws['A5'] = f"Profit: {context['total_profit']:.2f}"

        headers = ['Product', 'Quantity', 'Revenue', 'Cost', 'Profit', 'Margin %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row = 8
        for product_name, item in context['products_profit']:
            ws.cell(row=row, column=1).value = product_name
            ws.cell(row=row, column=2).value = int(item['quantity'])
            ws.cell(row=row, column=3).value = float(item['revenue'])
            ws.cell(row=row, column=4).value = float(item['cost'])
            ws.cell(row=row, column=5).value = float(item['profit'])
            ws.cell(row=row, column=6).value = float(item['margin'])
            row += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="profit_report_{start_date}_{end_date}.xlsx"'
        wb.save(response)
        return response
    except ImportError as e:
        messages.error(request, f'Excel generation not available. Please install openpyxl: {str(e)}')
        return redirect('profit_report')
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('profit_report')


@login_required
def print_income_statement_html(request):
    default_start_date, default_end_date = _get_current_week_date_range()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = default_start_date
    else:
        try:
            start_date = datetime.fromisoformat(start_date).date()
        except ValueError:
            start_date = default_start_date

    if not end_date:
        end_date = default_end_date
    else:
        try:
            end_date = datetime.fromisoformat(end_date).date()
        except ValueError:
            end_date = default_end_date

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    context = _build_income_statement_context(start_date, end_date)
    context['now'] = timezone.now()
    return render(request, 'inventory/print_income_statement.html', context)


@login_required
def print_income_statement_pdf(request):
    default_start_date, default_end_date = _get_current_week_date_range()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = default_start_date
    else:
        try:
            start_date = datetime.fromisoformat(start_date).date()
        except ValueError:
            start_date = default_start_date

    if not end_date:
        end_date = default_end_date
    else:
        try:
            end_date = datetime.fromisoformat(end_date).date()
        except ValueError:
            end_date = default_end_date

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    context = _build_income_statement_context(start_date, end_date)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="income_statement_{start_date}_{end_date}.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.9 * inch, height=0.9 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.12 * inch))

        title = Paragraph(
            f"<b>Income Statement - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        summary_lines = [
            ['Sales Revenue', f"Rs.{context['sales_revenue']:.2f}"],
            ['Cost of Goods Sold (COGS)', f"Rs.{context['cost_of_goods_sold']:.2f}"],
            ['Gross Profit', f"Rs.{context['gross_profit']:.2f}"],
            ['Other Operating Income', f"Rs.{context['operating_income_total']:.2f}"],
            ['Operating Expenses', f"Rs.{context['operating_expense_total']:.2f}"],
            ['Net Profit (Before Tax)', f"Rs.{context['net_profit_before_tax']:.2f}"],
        ]
        summary_table = Table(summary_lines, colWidths=[3.4 * inch, 2.0 * inch])
        summary_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 2), (-1, 2), colors.lightgrey),
            ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph('<b>Sales Breakdown (By Product)</b>', styles['Heading3']))
        data = [['Product', 'Qty', 'Revenue', 'COGS', 'Gross Profit']]
        for row in context['sales_breakdown']:
            data.append([
                row['product_name'],
                str(row['quantity']),
                f"Rs.{row['revenue']:.2f}",
                f"Rs.{row['cost']:.2f}",
                f"Rs.{row['gross_profit']:.2f}",
            ])

        table = Table(data, colWidths=[2.2 * inch, 0.7 * inch, 1.0 * inch, 1.0 * inch, 1.1 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        return response
    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('income_statement')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('income_statement')


@login_required
def print_income_statement_excel(request):
    default_start_date, default_end_date = _get_current_week_date_range()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date:
        start_date = default_start_date
    else:
        try:
            start_date = datetime.fromisoformat(start_date).date()
        except ValueError:
            start_date = default_start_date

    if not end_date:
        end_date = default_end_date
    else:
        try:
            end_date = datetime.fromisoformat(end_date).date()
        except ValueError:
            end_date = default_end_date

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    context = _build_income_statement_context(start_date, end_date)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'Income Statement'

        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16

        ws['A1'] = f"INCOME STATEMENT - {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = 'Sales Revenue'
        ws['B3'] = float(context['sales_revenue'])
        ws['A4'] = 'Cost of Goods Sold (COGS)'
        ws['B4'] = float(context['cost_of_goods_sold'])
        ws['A5'] = 'Gross Profit'
        ws['B5'] = float(context['gross_profit'])
        ws['A6'] = 'Other Operating Income'
        ws['B6'] = float(context['operating_income_total'])
        ws['A7'] = 'Operating Expenses'
        ws['B7'] = float(context['operating_expense_total'])
        ws['A8'] = 'Net Profit (Before Tax)'
        ws['B8'] = float(context['net_profit_before_tax'])

        for cell_ref in ['A5', 'B5', 'A8', 'B8']:
            ws[cell_ref].font = Font(bold=True)

        ws['A10'] = 'Sales Breakdown (By Product)'
        ws['A10'].font = Font(bold=True)

        headers = ['Product', 'Qty Sold', 'Revenue', 'COGS', 'Gross Profit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row_idx = 12
        for row in context['sales_breakdown']:
            ws.cell(row=row_idx, column=1).value = row['product_name']
            ws.cell(row=row_idx, column=2).value = int(row['quantity'])
            ws.cell(row=row_idx, column=3).value = float(row['revenue'])
            ws.cell(row=row_idx, column=4).value = float(row['cost'])
            ws.cell(row=row_idx, column=5).value = float(row['gross_profit'])
            row_idx += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="income_statement_{start_date}_{end_date}.xlsx"'
        )
        wb.save(response)
        return response
    except ImportError as e:
        messages.error(request, f'Excel generation not available. Please install openpyxl: {str(e)}')
        return redirect('income_statement')
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('income_statement')

# ==================== USER MANAGEMENT ====================

@login_required
@permission_required('auth.change_user', raise_exception=True)
def user_list(request):
    """List all users"""
    users = User.objects.all()
    context = {'users': users}
    return render(request, 'inventory/user_list.html', context)

@login_required
@permission_required('auth.add_user', raise_exception=True)
def add_user(request):
    """Add new user"""
    if request.method == 'POST':
        form = UserManagementForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'User created successfully')
            return redirect('user_list')
    else:
        form = UserManagementForm()
    
    context = {'form': form, 'title': 'Add New User'}
    return render(request, 'inventory/user_form.html', context)

@login_required
@permission_required('auth.change_user', raise_exception=True)
def edit_user(request, user_id):
    """Edit user"""
    user = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        form = UserManagementForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully')
            return redirect('user_list')
    else:
        form = UserManagementForm(instance=user)
    
    context = {'form': form, 'user': user, 'title': f'Edit {user.username}'}
    return render(request, 'inventory/user_form.html', context)

@login_required
@permission_required('auth.delete_user', raise_exception=True)
def delete_user(request, user_id):
    """Delete user"""
    user = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        user.delete()
        messages.success(request, f'User {user.username} deleted')
        return redirect('user_list')
    
        messages.success(request, f'User {user.username} deleted')
        return redirect('user_list')
    
    context = {'user': user}
    return render(request, 'inventory/delete_confirm.html', context)

# ==================== PRINT MODULE ====================

def _build_daily_data_sheet_print_context(request):
    """Build print context for Daily Sales Sheet with stock taken and balance columns."""
    selected_sales_date_raw = request.GET.get('sales_date') or request.GET.get('date')
    if selected_sales_date_raw:
        try:
            selected_sales_date = datetime.strptime(selected_sales_date_raw, '%Y-%m-%d').date()
        except ValueError:
            selected_sales_date = timezone.now().date()
    else:
        selected_sales_date = timezone.now().date()

    target_user = request.user
    if request.user.is_staff:
        salesperson_id_raw = (request.GET.get('salesperson') or '').strip()
        if salesperson_id_raw:
            try:
                target_user = User.objects.get(pk=int(salesperson_id_raw), is_active=True)
            except (User.DoesNotExist, ValueError, TypeError):
                target_user = request.user

    products = build_grouped_products_for_sales_date(selected_sales_date)

    current_grouped_products = group_active_products_by_name()
    for product in products:
        candidates = current_grouped_products.get(product['key'], [])
        product['stock'] = sum(max(0, item.current_stock) for item in candidates)

    stock_taken_map = {
        item.product_key: item.stock_taken_count
        for item in SalesStockTaken.objects.filter(
            salesperson=target_user,
            sales_date=selected_sales_date,
        )
    }

    sales_count_map = defaultdict(int)
    user_sales_qs = Sales.objects.filter(
        sale_date=selected_sales_date,
        recorded_by=target_user,
    ).select_related('product')
    for sale in user_sales_qs:
        key = normalize_sales_product_name(sale.product.name).lower()
        sales_count_map[key] += sale.quantity

    total_stock_taken_for_date = 0
    total_sales_count = 0
    total_estimated_total = Decimal('0.0')

    for product in products:
        stock_taken_count = stock_taken_map.get(product['key'], 0)
        sales_count = sales_count_map.get(product['key'], 0)
        balance_count = stock_taken_count - sales_count
        estimated_total = product['avg_price'] * Decimal(sales_count)

        product['stock_taken_count'] = stock_taken_count
        product['sales_count'] = sales_count
        product['balance_count'] = balance_count
        product['estimated_total'] = estimated_total

        total_stock_taken_for_date += stock_taken_count
        total_sales_count += sales_count
        total_estimated_total += estimated_total

    return {
        'selected_sales_date': selected_sales_date,
        'target_user': target_user,
        'products': products,
        'total_stock_taken_for_date': total_stock_taken_for_date,
        'total_sales_count': total_sales_count,
        'total_estimated_total': total_estimated_total,
    }


@login_required
def print_daily_data_sheet_html(request):
    """Print-friendly Daily Data Sheet view for salespersons."""
    context = _build_daily_data_sheet_print_context(request)
    context['now'] = timezone.now()
    return render(request, 'inventory/print_daily_data_sheet.html', context)

@login_required
def print_sales_html(request):
    """Print sales in HTML format (browser print-friendly)"""
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    selected_salesperson_id = request.GET.get('salesperson', '').strip()

    if isinstance(selected_date, str):
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()

    salesperson_filter = None
    if not request.user.is_staff:
        salesperson_filter = request.user
    elif selected_salesperson_id:
        try:
            salesperson_filter = User.objects.get(pk=int(selected_salesperson_id))
        except (ValueError, User.DoesNotExist):
            selected_salesperson_id = ''

    # Get sales for the selected date and group by normalized product name.
    daily_sales_qs = Sales.objects.filter(sale_date=selected_date)
    if salesperson_filter:
        daily_sales_qs = daily_sales_qs.filter(recorded_by=salesperson_filter)
    daily_sales_qs = daily_sales_qs.select_related('product', 'recorded_by').order_by('product__sku')
    daily_sales = build_sales_groups(daily_sales_qs)

    # Calculate totals
    total_sales_count = len(daily_sales)
    total_revenue = daily_sales_qs.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']
    total_quantity = daily_sales_qs.aggregate(
        total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
    )['total']

    context = {
        'selected_date': selected_date,
        'daily_sales': daily_sales,
        'total_sales_count': total_sales_count,
        'total_revenue': total_revenue,
        'total_quantity': total_quantity,
        'selected_salesperson': salesperson_filter,
        'now': timezone.now(),
    }

    return render(request, 'inventory/print_sales.html', context)

@login_required
def print_sales_pdf(request):
    """Export sales as PDF"""
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    selected_salesperson_id = request.GET.get('salesperson', '').strip()

    if isinstance(selected_date, str):
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()

    salesperson_filter = None
    if not request.user.is_staff:
        salesperson_filter = request.user
    elif selected_salesperson_id:
        try:
            salesperson_filter = User.objects.get(pk=int(selected_salesperson_id))
        except (ValueError, User.DoesNotExist):
            selected_salesperson_id = ''

    # Get sales for the selected date and group by normalized product name.
    daily_sales_qs = Sales.objects.filter(sale_date=selected_date)
    if salesperson_filter:
        daily_sales_qs = daily_sales_qs.filter(recorded_by=salesperson_filter)
    daily_sales_qs = daily_sales_qs.select_related('product', 'recorded_by').order_by('product__sku')
    daily_sales = build_sales_groups(daily_sales_qs)

    # Calculate totals
    total_revenue = daily_sales_qs.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']
    total_quantity = daily_sales_qs.aggregate(
        total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
    )['total']

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales_{selected_date}.pdf"'
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        logo_path = _get_report_logo_path()
        if logo_path:
            elements.append(Image(logo_path, width=0.9 * inch, height=0.9 * inch, hAlign='CENTER'))
            elements.append(Spacer(1, 0.12 * inch))
        
        # Title
        title = Paragraph(f"<b>Sales Report - {selected_date.strftime('%d %B %Y')}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary
        if salesperson_filter:
            salesperson_name = salesperson_filter.get_full_name() or salesperson_filter.username
            summary_text = f"<b>Summary:</b> Salesperson: {salesperson_name} | Total Sales: {len(daily_sales)} | Total Quantity: {total_quantity} | Total Revenue: ₹{total_revenue:,.2f}"
        else:
            summary_text = f"<b>Summary:</b> Total Sales: {len(daily_sales)} | Total Quantity: {total_quantity} | Total Revenue: ₹{total_revenue:,.2f}"
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Data for table
        data = [['Product Name', 'Qty', 'Unit Price', 'Total Price', 'Recorded By']]
        for sale in daily_sales:
            data.append([
                sale['product_name'],
                str(sale['quantity']),
                f"₹{sale['unit_price']:.2f}",
                f"₹{sale['total_price']:.2f}",
                sale['recorded_by']
            ])
        
        # Add totals row
        data.append(['TOTAL', str(total_quantity), '', f"₹{total_revenue:.2f}", ''])
        
        # Create table
        table = Table(data, colWidths=[2.2*inch, 0.8*inch, 1.0*inch, 1.1*inch, 1.6*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return response
    except ImportError as e:
        messages.error(request, f'PDF generation not available. Please install reportlab: {str(e)}')
        return redirect('view_sales')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('view_sales')


@login_required
def print_sales_jpeg(request):
    """Export sales as JPEG image."""
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    selected_salesperson_id = request.GET.get('salesperson', '').strip()

    if isinstance(selected_date, str):
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()

    salesperson_filter = None
    if not request.user.is_staff:
        salesperson_filter = request.user
    elif selected_salesperson_id:
        try:
            salesperson_filter = User.objects.get(pk=int(selected_salesperson_id))
        except (ValueError, User.DoesNotExist):
            selected_salesperson_id = ''

    daily_sales_qs = Sales.objects.filter(sale_date=selected_date)
    if salesperson_filter:
        daily_sales_qs = daily_sales_qs.filter(recorded_by=salesperson_filter)
    daily_sales_qs = daily_sales_qs.select_related('product', 'recorded_by').order_by('product__sku')
    daily_sales = build_sales_groups(daily_sales_qs)

    total_quantity = daily_sales_qs.aggregate(
        total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
    )['total']
    total_revenue = daily_sales_qs.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']

    redirect_target = 'view_sales' if request.user.is_staff else 'quick_sales_entry'

    try:
        from PIL import Image, ImageDraw, ImageFont

        left = 40
        row_h = 34
        img_width = 1700
        header_top = 30
        table_top = 230
        cols = [620, 130, 190, 210, 470]
        table_width = sum(cols)

        row_count = max(1, len(daily_sales))
        img_height = table_top + ((row_count + 2) * row_h) + 60

        image = Image.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(image)

        title_font = ImageFont.load_default()
        body_font = ImageFont.load_default()

        draw.text((left, header_top), 'DAILY SALES SHEET', fill=(87, 40, 9), font=title_font)
        draw.text((left, header_top + 30), f'Date: {selected_date.strftime("%Y-%m-%d")}', fill=(50, 50, 50), font=body_font)

        if salesperson_filter:
            salesperson_name = salesperson_filter.get_full_name() or salesperson_filter.username
            draw.text((left, header_top + 55), f'Salesperson: {salesperson_name}', fill=(50, 50, 50), font=body_font)

        summary_text = (
            f'Total Sales: {len(daily_sales)}    '
            f'Total Quantity: {total_quantity}    '
            f'Total Revenue: Rs.{total_revenue:,.2f}'
        )
        draw.text((left, header_top + 85), summary_text, fill=(40, 40, 40), font=body_font)

        headers = ['Product Name', 'Qty', 'Unit Price', 'Total Price', 'Recorded By']

        x = left
        for i, header in enumerate(headers):
            draw.rectangle([(x, table_top), (x + cols[i], table_top + row_h)], fill=(87, 40, 9))
            draw.text((x + 8, table_top + 10), header, fill='white', font=body_font)
            x += cols[i]

        current_y = table_top + row_h
        if daily_sales:
            for sale in daily_sales:
                values = [
                    sale['product_name'],
                    str(sale['quantity']),
                    f"Rs.{sale['unit_price']:.2f}",
                    f"Rs.{sale['total_price']:.2f}",
                    sale['recorded_by'],
                ]
                x = left
                for i, value in enumerate(values):
                    draw.rectangle([(x, current_y), (x + cols[i], current_y + row_h)], outline=(180, 180, 180), width=1)
                    draw.text((x + 8, current_y + 10), str(value), fill=(30, 30, 30), font=body_font)
                    x += cols[i]
                current_y += row_h
        else:
            draw.rectangle([(left, current_y), (left + table_width, current_y + row_h)], outline=(180, 180, 180), width=1)
            draw.text((left + 8, current_y + 10), 'No sales found for selected date.', fill=(120, 120, 120), font=body_font)
            current_y += row_h

        totals_values = ['TOTAL', str(total_quantity), '', f"Rs.{total_revenue:.2f}", '']
        x = left
        for i, value in enumerate(totals_values):
            draw.rectangle([(x, current_y), (x + cols[i], current_y + row_h)], fill=(242, 232, 219), outline=(170, 170, 170), width=1)
            draw.text((x + 8, current_y + 10), str(value), fill=(40, 40, 40), font=body_font)
            x += cols[i]

        image_bytes = io.BytesIO()
        image.save(image_bytes, format='JPEG', quality=92)
        image_bytes.seek(0)

        response = HttpResponse(image_bytes.getvalue(), content_type='image/jpeg')
        response['Content-Disposition'] = f'attachment; filename="sales_{selected_date}.jpg"'
        return response

    except ImportError as e:
        messages.error(request, f'JPEG generation not available. Please install Pillow: {str(e)}')
        return redirect(redirect_target)
    except Exception as e:
        messages.error(request, f'Error generating JPEG: {str(e)}')
        return redirect(redirect_target)

@login_required
def print_sales_excel(request):
    """Export sales as Excel"""
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    selected_salesperson_id = request.GET.get('salesperson', '').strip()

    if isinstance(selected_date, str):
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()

    salesperson_filter = None
    if not request.user.is_staff:
        salesperson_filter = request.user
    elif selected_salesperson_id:
        try:
            salesperson_filter = User.objects.get(pk=int(selected_salesperson_id))
        except (ValueError, User.DoesNotExist):
            selected_salesperson_id = ''

    # Get sales for the selected date and group by normalized product name.
    daily_sales_qs = Sales.objects.filter(sale_date=selected_date)
    if salesperson_filter:
        daily_sales_qs = daily_sales_qs.filter(recorded_by=salesperson_filter)
    daily_sales_qs = daily_sales_qs.select_related('product', 'recorded_by').order_by('product__sku')
    daily_sales = build_sales_groups(daily_sales_qs)

    # Calculate totals
    total_quantity = daily_sales_qs.aggregate(
        total=Coalesce(Sum('quantity'), 0, output_field=DecimalField())
    )['total']
    total_revenue = daily_sales_qs.aggregate(
        total=Coalesce(Sum('total_price'), 0, output_field=DecimalField())
    )['total']

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Set column widths
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 24
        
        # Title
        ws['A1'] = f"SALES REPORT - {selected_date.strftime('%d %B %Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Summary section
        ws['A3'] = 'Summary:'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = f"Total Sales Count: {len(daily_sales)}"
        ws['A5'] = f"Total Quantity: {total_quantity}"
        ws['A6'] = f"Total Revenue: ₹{total_revenue:,.2f}"
        if salesperson_filter:
            ws['A7'] = f"Salesperson: {salesperson_filter.get_full_name() or salesperson_filter.username}"
        
        # Headers
        headers = ['Product Name', 'Quantity', 'Unit Price', 'Total Price', 'Recorded By']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 9
        for sale in daily_sales:
            ws.cell(row=row, column=1).value = sale['product_name']
            ws.cell(row=row, column=2).value = sale['quantity']
            ws.cell(row=row, column=3).value = float(sale['unit_price'])
            ws.cell(row=row, column=4).value = float(sale['total_price'])
            ws.cell(row=row, column=5).value = sale['recorded_by']
            
            # Center align numeric columns
            for col in [2, 3, 4]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Totals row
        totals_row = row
        ws.cell(row=totals_row, column=1).value = "TOTAL"
        ws.cell(row=totals_row, column=1).font = Font(bold=True)
        ws.cell(row=totals_row, column=2).value = total_quantity
        ws.cell(row=totals_row, column=2).font = Font(bold=True)
        ws.cell(row=totals_row, column=4).value = float(total_revenue)
        ws.cell(row=totals_row, column=4).font = Font(bold=True)
        
        # Format currency columns
        for r in range(9, totals_row + 1):
            ws.cell(row=r, column=3).number_format = '₹#,##0.00'
            ws.cell(row=r, column=4).number_format = '₹#,##0.00'
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sales_{selected_date}.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('view_sales')


